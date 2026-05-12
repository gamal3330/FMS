from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import Iterable, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.audit import AuditLog
from app.models.database import DatabaseBackup, DatabaseJob
from app.models.enums import UserRole
from app.models.message import InternalMessage, InternalMessageAttachment, InternalMessageRecipient
from app.models.messaging_settings import MessageClassification, MessageType, MessagingSettings
from app.models.report import ReportExportLog, ReportTemplate, SavedReport, ScheduledReport
from app.models.request import Attachment, RequestApprovalStep, ServiceRequest
from app.models.settings import RequestTypeSetting, SettingsGeneral, SpecializedSection
from app.models.user import ActionPermission, Department, Role, User
from app.services.audit import write_audit
from app.services.messaging_settings_service import ensure_messaging_settings_schema
from app.services.pdf_fonts import register_arabic_pdf_font, rtl_text
from app.services.pdf_template import (
    draw_cover_header,
    draw_footer,
    draw_ltr_text,
    draw_page_header,
    draw_section_header,
    draw_text,
    format_pdf_datetime,
    pdf_theme,
)

router = APIRouter(prefix="/reports", tags=["Reports"])

FINAL_REQUEST_STATUSES = {"completed", "closed", "cancelled", "rejected"}
ADMIN_REPORT_ROLES = {UserRole.SUPER_ADMIN, UserRole.EXECUTIVE, UserRole.DEPARTMENT_MANAGER}
SENSITIVE_REPORT_TYPES = {"audit", "users-permissions"}
DANGEROUS_ACTIONS = {
    "reset_database",
    "restore_database",
    "apply_update",
    "manage_ai_settings",
    "audit_messages",
    "view_confidential_messages",
}

STATUS_LABELS = {
    "draft": "مسودة",
    "submitted": "مرسل",
    "pending_approval": "بانتظار الموافقة",
    "returned_for_edit": "معاد للتعديل",
    "approved": "معتمد",
    "rejected": "مرفوض",
    "in_implementation": "قيد التنفيذ",
    "in_progress": "قيد التنفيذ",
    "completed": "مكتمل",
    "closed": "مغلق",
    "cancelled": "ملغي",
    "reopened": "معاد فتحه",
}

PRIORITY_LABELS = {
    "low": "منخفضة",
    "normal": "عادية",
    "medium": "متوسطة",
    "high": "مرتفعة",
    "urgent": "عاجلة",
    "critical": "حرجة",
}

REPORT_TYPE_LABELS = {
    "summary": "نظرة عامة",
    "requests": "تقارير الطلبات",
    "approvals": "تقارير الموافقات",
    "sla": "تقارير SLA والتأخير",
    "users-permissions": "تقارير المستخدمين والصلاحيات",
    "messaging": "تقارير المراسلات",
    "attachments": "تقارير المرفقات",
    "audit": "تقارير التدقيق",
}

EXCEL_HEADERS = [
    "رقم الطلب",
    "العنوان",
    "مقدم الطلب",
    "الإدارة",
    "نوع الطلب",
    "القسم المختص",
    "الموظف المنفذ",
    "الحالة",
    "الأولوية",
    "تاريخ الإنشاء",
    "تاريخ الإغلاق",
    "حالة SLA",
]


class ReportFilters(BaseModel):
    date_from: date | None = None
    date_to: date | None = None
    department_id: int | None = None
    request_type_id: int | None = None
    status: str | None = None
    priority: str | None = None
    specialized_section_id: int | None = None
    requester_id: int | None = None
    assigned_user_id: int | None = None
    approval_step: str | None = None
    sla_status: str | None = None
    message_type: str | None = None
    audit_action: str | None = None

    @field_validator("status", "priority", "approval_step", "sla_status", "message_type", "audit_action")
    @classmethod
    def blank_to_none(cls, value: str | None) -> str | None:
        value = (value or "").strip()
        return value or None

    def compact(self) -> dict:
        data = self.model_dump()
        return {key: value.isoformat() if isinstance(value, date) else value for key, value in data.items() if value not in (None, "")}


class SavedReportPayload(BaseModel):
    name: str = Field(min_length=1, max_length=160)
    description: str | None = None
    report_type: str = Field(min_length=1, max_length=80)
    filters_json: dict = Field(default_factory=dict)
    is_favorite: bool = False


class ReportTemplatePayload(BaseModel):
    name_ar: str = Field(min_length=1, max_length=160)
    code: str = Field(min_length=1, max_length=100)
    report_type: str = Field(min_length=1, max_length=80)
    description: str | None = None
    default_filters_json: dict = Field(default_factory=dict)
    default_columns_json: list = Field(default_factory=list)
    is_active: bool = True


class ScheduledReportPayload(BaseModel):
    name: str = Field(min_length=1, max_length=160)
    report_template_id: int | None = None
    frequency: Literal["daily", "weekly", "monthly"] = "monthly"
    run_time: str = Field(default="08:00", pattern=r"^\d{2}:\d{2}$")
    recipients_json: list = Field(default_factory=list)
    export_format: Literal["excel", "pdf"] = "excel"
    is_active: bool = True


def get_report_filters(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    request_type_id: int | None = Query(default=None),
    status_value: str | None = Query(default=None, alias="status"),
    priority: str | None = Query(default=None),
    specialized_section_id: int | None = Query(default=None),
    requester_id: int | None = Query(default=None),
    assigned_user_id: int | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    approval_step: str | None = Query(default=None),
    sla_status: str | None = Query(default=None),
    message_type: str | None = Query(default=None),
    audit_action: str | None = Query(default=None),
) -> ReportFilters:
    filters = ReportFilters(
        date_from=date_from or from_date,
        date_to=date_to or to_date,
        department_id=department_id,
        request_type_id=request_type_id,
        status=status_value,
        priority=priority,
        specialized_section_id=specialized_section_id,
        requester_id=requester_id or employee_id,
        assigned_user_id=assigned_user_id,
        approval_step=approval_step,
        sla_status=sla_status,
        message_type=message_type,
        audit_action=audit_action,
    )
    if filters.date_from and filters.date_to and filters.date_from > filters.date_to:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
    return filters


def value_of(value) -> str:
    return str(getattr(value, "value", value or ""))


def label(value, labels: dict[str, str]) -> str:
    return labels.get(value_of(value), value_of(value))


def datetime_label(value: datetime | None) -> str:
    if not isinstance(value, datetime):
        return ""
    return value.strftime("%Y/%m/%d %H:%M")


def rtl(value: object) -> str:
    return rtl_text(value)


def request_type_label(item: ServiceRequest, type_map: dict[int, RequestTypeSetting] | None = None) -> str:
    snapshot = getattr(item, "request_type_snapshot", None) or {}
    form_data = getattr(item, "form_data", None) or {}
    request_type_id = getattr(item, "request_type_id", None)
    request_type = type_map.get(request_type_id) if type_map and request_type_id else None
    return (
        str(snapshot.get("name_ar") or "")
        or str(form_data.get("request_type_label") or "")
        or (request_type.name_ar if request_type else "")
        or str(form_data.get("request_type_code") or "")
        or value_of(item.request_type)
    )


def specialized_section_label(item: ServiceRequest, type_map: dict[int, RequestTypeSetting] | None = None) -> str:
    snapshot = getattr(item, "request_type_snapshot", None) or {}
    form_data = getattr(item, "form_data", None) or {}
    request_type_id = getattr(item, "request_type_id", None)
    request_type = type_map.get(request_type_id) if type_map and request_type_id else None
    return (
        str(snapshot.get("specialized_section_name") or "")
        or str(snapshot.get("assigned_section_label") or "")
        or str(form_data.get("assigned_section_label") or "")
        or (request_type.assigned_section if request_type else "")
        or ""
    )


def sla_status(item: ServiceRequest) -> str:
    due = getattr(item, "sla_due_at", None)
    if not due:
        return "no_sla"
    status_value = value_of(getattr(item, "status", None))
    now = datetime.now(timezone.utc)
    normalized_due = due if due.tzinfo else due.replace(tzinfo=timezone.utc)
    if status_value in FINAL_REQUEST_STATUSES:
        closed_at_value = getattr(item, "closed_at", None)
        if closed_at_value:
            closed_at = closed_at_value if closed_at_value.tzinfo else closed_at_value.replace(tzinfo=timezone.utc)
            return "met" if closed_at <= normalized_due else "breached"
        return "met"
    return "breached" if normalized_due < now else "on_track"


def sla_label(value: str) -> str:
    return {
        "met": "ملتزم",
        "breached": "متأخر",
        "on_track": "ضمن الوقت",
        "no_sla": "بدون SLA",
        "near_breach": "قريب من التأخير",
    }.get(value, value)


def ensure_sensitive_report_allowed(actor: User, report_type: str) -> None:
    if report_type not in SENSITIVE_REPORT_TYPES:
        return
    if actor.role not in ADMIN_REPORT_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="لا تملك صلاحية عرض هذا التقرير")


def request_scope_condition(db: Session, actor: User):
    role = actor.role
    if role in {UserRole.SUPER_ADMIN, UserRole.EXECUTIVE}:
        return None

    conditions = [ServiceRequest.requester_id == actor.id]
    if role == UserRole.DIRECT_MANAGER:
        conditions.append(ServiceRequest.requester_id.in_(select(User.id).where(User.manager_id == actor.id)))
    if role == UserRole.DEPARTMENT_MANAGER:
        managed_departments = select(Department.id).where(Department.manager_id == actor.id, Department.is_active == True)
        if actor.department_id:
            conditions.append(ServiceRequest.department_id == actor.department_id)
        conditions.append(ServiceRequest.department_id.in_(managed_departments))
    if role == UserRole.IT_STAFF:
        conditions.append(ServiceRequest.assigned_to_id == actor.id)
        if actor.specialized_section_id:
            section = db.get(SpecializedSection, actor.specialized_section_id)
            if section:
                request_type_ids = select(RequestTypeSetting.id).where(
                    RequestTypeSetting.is_active == True,
                    or_(
                        RequestTypeSetting.assigned_section == section.code,
                        RequestTypeSetting.assigned_department_id == section.department_id,
                    ),
                )
                conditions.append(ServiceRequest.request_type_id.in_(request_type_ids))
    return or_(*conditions)


def apply_request_scope(stmt, db: Session, actor: User):
    condition = request_scope_condition(db, actor)
    return stmt.where(condition) if condition is not None else stmt


def apply_request_filters(stmt, filters: ReportFilters, db: Session):
    if filters.date_from:
        stmt = stmt.where(ServiceRequest.created_at >= datetime.combine(filters.date_from, time.min))
    if filters.date_to:
        stmt = stmt.where(ServiceRequest.created_at <= datetime.combine(filters.date_to, time.max))
    if filters.department_id:
        stmt = stmt.where(ServiceRequest.department_id == filters.department_id)
    if filters.request_type_id:
        stmt = stmt.where(ServiceRequest.request_type_id == filters.request_type_id)
    if filters.status:
        stmt = stmt.where(ServiceRequest.status == filters.status)
    if filters.priority:
        stmt = stmt.where(ServiceRequest.priority == filters.priority)
    if filters.requester_id:
        stmt = stmt.where(ServiceRequest.requester_id == filters.requester_id)
    if filters.assigned_user_id:
        stmt = stmt.where(ServiceRequest.assigned_to_id == filters.assigned_user_id)
    if filters.specialized_section_id:
        section = db.get(SpecializedSection, filters.specialized_section_id)
        if section:
            request_type_ids = select(RequestTypeSetting.id).where(
                or_(
                    RequestTypeSetting.assigned_section == section.code,
                    RequestTypeSetting.assigned_department_id == section.department_id,
                )
            )
            stmt = stmt.where(ServiceRequest.request_type_id.in_(request_type_ids))
        else:
            stmt = stmt.where(ServiceRequest.request_type_id == -1)
    now = datetime.now(timezone.utc)
    if filters.sla_status == "breached":
        stmt = stmt.where(ServiceRequest.sla_due_at.is_not(None), ServiceRequest.sla_due_at < now, ServiceRequest.status.not_in(FINAL_REQUEST_STATUSES))
    elif filters.sla_status == "met":
        stmt = stmt.where(ServiceRequest.sla_due_at.is_not(None), ServiceRequest.closed_at.is_not(None), ServiceRequest.closed_at <= ServiceRequest.sla_due_at)
    elif filters.sla_status == "on_track":
        stmt = stmt.where(ServiceRequest.sla_due_at.is_not(None), ServiceRequest.sla_due_at >= now, ServiceRequest.status.not_in(FINAL_REQUEST_STATUSES))
    elif filters.sla_status == "no_sla":
        stmt = stmt.where(ServiceRequest.sla_due_at.is_(None))
    return stmt


def request_statement(db: Session, actor: User, filters: ReportFilters, with_options: bool = False, ordered: bool = True):
    stmt = select(ServiceRequest)
    if with_options:
        stmt = stmt.options(
            selectinload(ServiceRequest.requester),
            selectinload(ServiceRequest.department),
            selectinload(ServiceRequest.assigned_to),
            selectinload(ServiceRequest.attachments),
        )
    stmt = apply_request_filters(stmt, filters, db)
    stmt = apply_request_scope(stmt, db, actor)
    if ordered:
        stmt = stmt.order_by(ServiceRequest.created_at.desc())
    return stmt


def total_for(db: Session, stmt) -> int:
    return int(db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery())) or 0)


def request_type_map(db: Session, request_type_ids: Iterable[int | None]) -> dict[int, RequestTypeSetting]:
    ids = sorted({item for item in request_type_ids if item})
    if not ids:
        return {}
    return {item.id: item for item in db.scalars(select(RequestTypeSetting).where(RequestTypeSetting.id.in_(ids))).all()}


def requester_name(item: ServiceRequest) -> str:
    return item.requester.full_name_ar if getattr(item, "requester", None) else ""


def department_name(item: ServiceRequest) -> str:
    return item.department.name_ar if getattr(item, "department", None) else ""


def assigned_name(item: ServiceRequest) -> str:
    return item.assigned_to.full_name_ar if getattr(item, "assigned_to", None) else ""


def request_row(item: ServiceRequest, type_map: dict[int, RequestTypeSetting]) -> dict:
    created_at = item.created_at
    closed_at = item.closed_at
    duration_hours = None
    if isinstance(created_at, datetime) and isinstance(closed_at, datetime):
        duration_hours = round((closed_at - created_at).total_seconds() / 3600, 1)
    item_sla_status = sla_status(item)
    return {
        "id": item.id,
        "request_number": item.request_number,
        "title": item.title,
        "request_type": request_type_label(item, type_map),
        "request_type_id": item.request_type_id,
        "requester": requester_name(item),
        "requester_id": item.requester_id,
        "department": department_name(item),
        "department_id": item.department_id,
        "specialized_section": specialized_section_label(item, type_map),
        "assigned_user": assigned_name(item),
        "assigned_user_id": item.assigned_to_id,
        "status": value_of(item.status),
        "status_label": label(item.status, STATUS_LABELS),
        "priority": value_of(item.priority),
        "priority_label": label(item.priority, PRIORITY_LABELS),
        "created_at": item.created_at.isoformat() if isinstance(item.created_at, datetime) else None,
        "closed_at": item.closed_at.isoformat() if isinstance(item.closed_at, datetime) else None,
        "duration_hours": duration_hours,
        "sla_due_at": item.sla_due_at.isoformat() if isinstance(item.sla_due_at, datetime) else None,
        "sla_status": item_sla_status,
        "sla_status_label": sla_label(item_sla_status),
        "attachments_count": len(getattr(item, "attachments", []) or []),
    }


def chart_from_counts(rows, label_key: str = "label") -> list[dict]:
    return [{label_key: str(name or "غير محدد"), "count": int(count or 0)} for name, count in rows]


def relabel_chart(rows: list[dict], labels: dict[str, str]) -> list[dict]:
    return [{**row, "label": labels.get(str(row.get("label")), str(row.get("label") or "غير محدد"))} for row in rows]


def grouped_request_counts(db: Session, actor: User, filters: ReportFilters, column) -> list[dict]:
    stmt = select(column, func.count(ServiceRequest.id))
    stmt = apply_request_filters(stmt, filters, db)
    stmt = apply_request_scope(stmt, db, actor).group_by(column)
    return chart_from_counts(db.execute(stmt).all())


def request_department_counts(db: Session, actor: User, filters: ReportFilters) -> list[dict]:
    stmt = select(Department.name_ar, func.count(ServiceRequest.id)).join(Department, Department.id == ServiceRequest.department_id, isouter=True)
    stmt = apply_request_filters(stmt, filters, db)
    stmt = apply_request_scope(stmt, db, actor).group_by(Department.name_ar)
    return chart_from_counts(db.execute(stmt).all())


def request_type_counts(db: Session, actor: User, filters: ReportFilters) -> list[dict]:
    stmt = select(RequestTypeSetting.name_ar, func.count(ServiceRequest.id)).join(RequestTypeSetting, RequestTypeSetting.id == ServiceRequest.request_type_id, isouter=True)
    stmt = apply_request_filters(stmt, filters, db)
    stmt = apply_request_scope(stmt, db, actor).group_by(RequestTypeSetting.name_ar)
    return chart_from_counts(db.execute(stmt).all())


def request_summary(db: Session, actor: User, filters: ReportFilters) -> dict:
    base = request_statement(db, actor, filters, ordered=False)
    total = total_for(db, base)
    completed = total_for(db, base.where(ServiceRequest.status.in_(["completed", "closed"])))
    open_count = total_for(db, base.where(ServiceRequest.status.not_in(FINAL_REQUEST_STATUSES)))
    breached = total_for(db, base.where(ServiceRequest.sla_due_at.is_not(None), ServiceRequest.sla_due_at < datetime.now(timezone.utc), ServiceRequest.status.not_in(FINAL_REQUEST_STATUSES)))
    with_sla = total_for(db, base.where(ServiceRequest.sla_due_at.is_not(None)))
    return {
        "total_requests": total,
        "completed_requests": completed,
        "open_requests": open_count,
        "delayed_requests": breached,
        "sla_compliance": round(((with_sla - breached) / with_sla) * 100, 1) if with_sla else 100,
        "average_completion_hours": average_completion_hours(db, base),
    }


def average_completion_hours(db: Session, base_stmt) -> float:
    items = db.scalars(
        base_stmt.where(ServiceRequest.closed_at.is_not(None))
        .order_by(None)
        .options()
        .limit(250)
    ).all()
    durations = []
    for item in items:
        if item.closed_at and item.created_at:
            durations.append((item.closed_at - item.created_at).total_seconds() / 3600)
    return round(sum(durations) / len(durations), 1) if durations else 0


def messaging_scope_condition(actor: User):
    if actor.role in {UserRole.SUPER_ADMIN, UserRole.EXECUTIVE}:
        return None
    return or_(InternalMessage.sender_id == actor.id, InternalMessageRecipient.recipient_id == actor.id)


def apply_message_filters(stmt, filters: ReportFilters):
    if filters.date_from:
        stmt = stmt.where(InternalMessage.created_at >= datetime.combine(filters.date_from, time.min))
    if filters.date_to:
        stmt = stmt.where(InternalMessage.created_at <= datetime.combine(filters.date_to, time.max))
    if filters.message_type:
        stmt = stmt.where(InternalMessage.message_type == filters.message_type)
    if filters.priority:
        stmt = stmt.where(InternalMessage.priority == filters.priority)
    if filters.request_type_id:
        stmt = stmt.join(ServiceRequest, ServiceRequest.id == InternalMessage.related_request_id, isouter=True).where(ServiceRequest.request_type_id == filters.request_type_id)
    return stmt


def apply_message_scope(stmt, actor: User):
    condition = messaging_scope_condition(actor)
    if condition is None:
        return stmt
    return stmt.join(InternalMessageRecipient, InternalMessageRecipient.message_id == InternalMessage.id, isouter=True).where(condition)


def serialize_saved_report(item: SavedReport) -> dict:
    return {
        "id": item.id,
        "name": item.name,
        "description": item.description,
        "report_type": item.report_type,
        "report_type_label": REPORT_TYPE_LABELS.get(item.report_type, item.report_type),
        "filters_json": item.filters_json or {},
        "is_favorite": item.is_favorite,
        "created_by": item.created_by,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


def serialize_template(item: ReportTemplate) -> dict:
    return {
        "id": item.id,
        "name_ar": item.name_ar,
        "code": item.code,
        "report_type": item.report_type,
        "report_type_label": REPORT_TYPE_LABELS.get(item.report_type, item.report_type),
        "description": item.description,
        "default_filters_json": item.default_filters_json or {},
        "default_columns_json": item.default_columns_json or [],
        "is_active": item.is_active,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


def serialize_schedule(item: ScheduledReport) -> dict:
    return {
        "id": item.id,
        "name": item.name,
        "report_template_id": item.report_template_id,
        "template_name": item.template.name_ar if item.template else "",
        "frequency": item.frequency,
        "run_time": item.run_time,
        "recipients_json": item.recipients_json or [],
        "export_format": item.export_format,
        "is_active": item.is_active,
        "last_run_at": item.last_run_at.isoformat() if item.last_run_at else None,
        "next_run_at": item.next_run_at.isoformat() if item.next_run_at else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


def ensure_default_templates(db: Session, actor: User | None = None) -> None:
    defaults = [
        ("تقرير الطلبات الشهري", "monthly_requests", "requests", "ملخص شهري لحركة الطلبات"),
        ("تقرير الطلبات المتأخرة", "delayed_requests", "sla", "طلبات متأخرة أو قريبة من التأخير"),
        ("تقرير أداء الأقسام المختصة", "section_performance", "requests", "حجم الطلبات حسب القسم المختص"),
        ("تقرير الموافقات", "approval_report", "approvals", "حركة الموافقات والانتظار"),
        ("تقرير SLA", "sla_report", "sla", "مؤشرات الالتزام بزمن الخدمة"),
        ("تقرير المراسلات الرسمية", "official_messages", "messaging", "المراسلات الرسمية والمرتبطة بالطلبات"),
        ("تقرير التدقيق", "audit_report", "audit", "سجل العمليات الحساسة"),
    ]
    existing = {row[0] for row in db.execute(select(ReportTemplate.code)).all()}
    for name, code, report_type, description in defaults:
        if code not in existing:
            db.add(ReportTemplate(name_ar=name, code=code, report_type=report_type, description=description, created_by=actor.id if actor else None))
    db.flush()


def log_report_action(db: Session, actor: User, action: str, report_type: str, filters: dict | None = None, request: Request | None = None) -> None:
    write_audit(
        db,
        action,
        "report",
        actor=actor,
        entity_id=report_type,
        metadata={"report_type": report_type, "filters": filters or {}},
        ip_address=request.client.host if request and request.client else None,
        user_agent=request.headers.get("user-agent") if request else None,
    )


@router.get("/summary")
def reports_summary(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
):
    summary = request_summary(db, actor, filters)
    message_count = messaging_count(db, actor, filters)
    last_export = db.scalar(select(ReportExportLog).where(ReportExportLog.exported_by == actor.id).order_by(ReportExportLog.exported_at.desc()).limit(1))
    data = {
        "cards": {
            **summary,
            "total_messages": message_count,
            "last_exported_report": last_export.exported_at.isoformat() if last_export else None,
        },
        "charts": {
            "requests_by_status": relabel_chart(grouped_request_counts(db, actor, filters, ServiceRequest.status), STATUS_LABELS),
            "requests_by_priority": relabel_chart(grouped_request_counts(db, actor, filters, ServiceRequest.priority), PRIORITY_LABELS),
            "requests_by_department": request_department_counts(db, actor, filters),
            "requests_by_type": request_type_counts(db, actor, filters),
            "requests_by_month": requests_by_month(db, actor, filters),
            "sla_trend": sla_trend(db, actor, filters),
        },
    }
    log_report_action(db, actor, "report_viewed", "summary", filters.compact(), request)
    db.commit()
    return data


@router.get("/requests")
def request_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    base = request_statement(db, actor, filters, with_options=True)
    total = total_for(db, base)
    items = db.scalars(base.offset((page - 1) * page_size).limit(page_size)).all()
    type_map = request_type_map(db, [item.request_type_id for item in items])
    log_report_action(db, actor, "report_viewed", "requests", filters.compact(), request)
    db.commit()
    return {
        "summary": request_summary(db, actor, filters),
        "items": [request_row(item, type_map) for item in items],
        "charts": {
            "by_status": relabel_chart(grouped_request_counts(db, actor, filters, ServiceRequest.status), STATUS_LABELS),
            "by_priority": relabel_chart(grouped_request_counts(db, actor, filters, ServiceRequest.priority), PRIORITY_LABELS),
            "by_request_type": request_type_counts(db, actor, filters),
            "over_time": requests_by_month(db, actor, filters),
        },
        "pagination": {"page": page, "page_size": page_size, "total": total},
    }


@router.get("/approvals")
def approval_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    request_ids = request_statement(db, actor, filters, ordered=False).with_only_columns(ServiceRequest.id).subquery()
    stmt = (
        select(RequestApprovalStep)
        .options(selectinload(RequestApprovalStep.request).selectinload(ServiceRequest.requester))
        .where(RequestApprovalStep.request_id.in_(select(request_ids.c.id)))
        .order_by(RequestApprovalStep.action_at.desc().nullslast(), RequestApprovalStep.sort_order.asc())
    )
    if filters.approval_step:
        needle = f"%{filters.approval_step}%"
        stmt = stmt.where(or_(RequestApprovalStep.step_type == filters.approval_step, RequestApprovalStep.step_name_ar.ilike(needle)))
    total = total_for(db, stmt)
    steps = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    users = user_map(db, [step.action_by for step in steps] + [step.approver_user_id for step in steps])
    rows = []
    for step in steps:
        req = step.request
        wait_hours = None
        if step.sla_due_at:
            ref = step.action_at or datetime.now(timezone.utc)
            wait_hours = round((ref - req.created_at).total_seconds() / 3600, 1) if req and req.created_at else None
        rows.append(
            {
                "request_id": req.id if req else None,
                "request_number": req.request_number if req else "",
                "request_type": request_type_label(req, {}) if req else "",
                "step_name": step.step_name_ar,
                "step_type": step.step_type,
                "approver": users.get(step.action_by or step.approver_user_id, ""),
                "status": step.status,
                "wait_hours": wait_hours,
                "action_at": step.action_at.isoformat() if step.action_at else None,
                "note": step.comments or "",
            }
        )
    pending = total_for(db, stmt.where(RequestApprovalStep.status.in_(["waiting", "pending"])))
    approved = total_for(db, stmt.where(RequestApprovalStep.status == "approved"))
    rejected = total_for(db, stmt.where(RequestApprovalStep.status == "rejected"))
    returned = total_for(db, stmt.where(RequestApprovalStep.status == "returned_for_edit"))
    log_report_action(db, actor, "report_viewed", "approvals", filters.compact(), request)
    db.commit()
    return {
        "summary": {
            "pending_approvals": pending,
            "approved_count": approved,
            "rejected_count": rejected,
            "returned_for_edit_count": returned,
            "average_approval_time": 0,
            "most_delayed_step": rows[0]["step_name"] if rows else "",
        },
        "items": rows,
        "charts": {"by_status": approval_status_counts(db, request_ids)},
        "pagination": {"page": page, "page_size": page_size, "total": total},
    }


@router.get("/sla")
def sla_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    base = request_statement(db, actor, filters, with_options=True)
    total = total_for(db, base)
    items = db.scalars(base.offset((page - 1) * page_size).limit(page_size)).all()
    type_map = request_type_map(db, [item.request_type_id for item in items])
    rows = []
    for item in items:
        row = request_row(item, type_map)
        delay_hours = None
        if row["sla_status"] == "breached" and item.sla_due_at:
            ref = item.closed_at or datetime.now(timezone.utc)
            normalized_due = item.sla_due_at if item.sla_due_at.tzinfo else item.sla_due_at.replace(tzinfo=timezone.utc)
            delay_hours = round((ref - normalized_due).total_seconds() / 3600, 1)
        row["delay_hours"] = delay_hours
        row["delay_reason"] = ""
        rows.append(row)
    summary = request_summary(db, actor, filters)
    log_report_action(db, actor, "report_viewed", "sla", filters.compact(), request)
    db.commit()
    return {
        "summary": {
            "sla_compliance": summary["sla_compliance"],
            "breached_requests": summary["delayed_requests"],
            "average_first_response_hours": 0,
            "average_resolution_hours": summary["average_completion_hours"],
            "requests_close_to_breach": 0,
            "requests_breached_today": breached_today(db, actor, filters),
        },
        "items": rows,
        "charts": {"sla_status": sla_status_counts(items), "over_time": sla_trend(db, actor, filters)},
        "pagination": {"page": page, "page_size": page_size, "total": total},
    }


@router.get("/users-permissions")
def users_permissions_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    ensure_sensitive_report_allowed(actor, "users-permissions")
    stmt = select(User).options(selectinload(User.department), selectinload(User.role_record)).order_by(User.full_name_ar)
    if filters.department_id:
        stmt = stmt.where(User.department_id == filters.department_id)
    total = total_for(db, stmt)
    users = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    dangerous_user_ids = set(
        row[0]
        for row in db.execute(
            select(ActionPermission.user_id).where(ActionPermission.user_id.is_not(None), ActionPermission.is_allowed == True, ActionPermission.action_code.in_(DANGEROUS_ACTIONS))
        ).all()
        if row[0]
    )
    rows = []
    for user in users:
        role_label = user.role_record.name_ar if user.role_record and user.role_record.name_ar else label(user.role, {
            "employee": "موظف",
            "direct_manager": "مدير مباشر",
            "it_staff": "مختص تنفيذ",
            "administration_manager": "مدير إدارة",
            "executive_management": "الإدارة التنفيذية",
            "super_admin": "مدير النظام",
        })
        high_privilege = user.role in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER, UserRole.EXECUTIVE} or user.id in dangerous_user_ids
        rows.append(
            {
                "id": user.id,
                "name": user.full_name_ar,
                "email": user.email,
                "department": user.department.name_ar if user.department else "",
                "role": role_label,
                "status": "نشط" if user.is_active else "غير نشط",
                "last_login": user.last_login_at.isoformat() if user.last_login_at else None,
                "has_high_privileges": high_privilege,
                "notes": "صلاحيات عالية" if high_privilege else "",
            }
        )
    log_report_action(db, actor, "report_viewed", "users-permissions", filters.compact(), request)
    db.commit()
    return {
        "summary": {
            "total_users": total_for(db, select(User)),
            "active_users": total_for(db, select(User).where(User.is_active == True)),
            "inactive_users": total_for(db, select(User).where(User.is_active == False)),
            "locked_users": total_for(db, select(User).where(User.is_locked == True)),
            "without_manager": total_for(db, select(User).where(User.manager_id.is_(None))),
            "without_department": total_for(db, select(User).where(User.department_id.is_(None))),
            "administrative_privileges": len([row for row in rows if row["has_high_privileges"]]),
        },
        "items": rows,
        "charts": {"by_role": user_role_counts(db), "by_department": user_department_counts(db)},
        "pagination": {"page": page, "page_size": page_size, "total": total},
    }


@router.get("/messaging")
def messaging_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    ensure_messaging_settings_schema(db)
    settings = db.scalar(select(MessagingSettings).limit(1))
    if settings and not settings.enable_messaging:
        return {"disabled": True, "message": "نظام المراسلات غير مفعل", "summary": {}, "items": [], "pagination": {"page": page, "page_size": page_size, "total": 0}}
    stmt = select(InternalMessage).options(selectinload(InternalMessage.sender), selectinload(InternalMessage.recipients).selectinload(InternalMessageRecipient.recipient), selectinload(InternalMessage.attachments))
    stmt = apply_message_filters(stmt, filters)
    stmt = apply_message_scope(stmt, actor).order_by(InternalMessage.created_at.desc()).distinct()
    total = total_for(db, stmt)
    messages = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    type_labels = {item.code: item.name_ar for item in db.scalars(select(MessageType)).all()}
    classification_labels = {item.code: item.name_ar for item in db.scalars(select(MessageClassification)).all()}
    rows = []
    for item in messages:
        rows.append(
            {
                "id": item.id,
                "message_uid": item.message_uid or f"MSG-{item.id}",
                "subject": item.subject,
                "message_type": item.message_type,
                "message_type_label": type_labels.get(item.message_type, item.message_type),
                "sender": item.sender.full_name_ar if item.sender else "",
                "recipients": "، ".join(recipient.recipient.full_name_ar for recipient in item.recipients if recipient.recipient),
                "related_request_id": item.related_request_id,
                "classification": classification_labels.get(item.classification_code, item.classification_code),
                "priority": item.priority,
                "priority_label": label(item.priority, PRIORITY_LABELS),
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "read_status": "مقروءة" if all(recipient.is_read for recipient in item.recipients) else "غير مقروءة",
                "attachments_count": len(item.attachments or []),
            }
        )
    log_report_action(db, actor, "report_viewed", "messaging", filters.compact(), request)
    db.commit()
    return {
        "summary": {
            "total_messages": total,
            "official_messages": message_count_by_type(db, actor, filters, "official_message"),
            "internal_messages": message_count_by_type(db, actor, filters, "internal_correspondence"),
            "clarification_requests": message_count_by_type(db, actor, filters, "clarification_request"),
            "unread_messages": unread_message_count(db, actor, filters),
            "linked_to_requests": total_for(db, apply_message_scope(apply_message_filters(select(InternalMessage).where(InternalMessage.related_request_id.is_not(None)), filters), actor)),
        },
        "items": rows,
        "charts": {"by_type": message_group_counts(db, actor, filters, InternalMessage.message_type)},
        "pagination": {"page": page, "page_size": page_size, "total": total},
    }


@router.get("/attachments")
def attachment_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    request_ids = request_statement(db, actor, filters, ordered=False).with_only_columns(ServiceRequest.id).subquery()
    request_attachments_stmt = (
        select(Attachment)
        .options(selectinload(Attachment.uploaded_by), selectinload(Attachment.request))
        .where(Attachment.request_id.in_(select(request_ids.c.id)))
        .order_by(Attachment.created_at.desc())
    )
    total_request_attachments = total_for(db, request_attachments_stmt)
    request_attachments = db.scalars(request_attachments_stmt.limit(page_size)).all()
    rows = [
        {
            "id": item.id,
            "file_name": item.original_name,
            "type": item.content_type,
            "size_bytes": item.size_bytes,
            "linked_to": item.request.request_number if item.request else "",
            "module": "الطلبات",
            "uploaded_by": item.uploaded_by.full_name_ar if item.uploaded_by else "",
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "downloads_count": 0,
            "status": "موجود",
        }
        for item in request_attachments
    ]
    log_report_action(db, actor, "report_viewed", "attachments", filters.compact(), request)
    db.commit()
    return {
        "summary": {
            "total_attachments": total_request_attachments,
            "total_storage_bytes": int(db.scalar(select(func.coalesce(func.sum(Attachment.size_bytes), 0)).where(Attachment.request_id.in_(select(request_ids.c.id)))) or 0),
            "large_files": len([row for row in rows if row["size_bytes"] > 10 * 1024 * 1024]),
            "missing_files": 0,
            "orphan_files": 0,
            "downloads": 0,
        },
        "items": rows,
        "charts": {"by_module": [{"label": "الطلبات", "count": total_request_attachments}]},
        "pagination": {"page": page, "page_size": page_size, "total": total_request_attachments},
    }


@router.get("/audit")
def audit_reports(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
):
    ensure_sensitive_report_allowed(actor, "audit")
    stmt = select(AuditLog).options(selectinload(AuditLog.actor)).order_by(AuditLog.created_at.desc())
    if filters.date_from:
        stmt = stmt.where(AuditLog.created_at >= datetime.combine(filters.date_from, time.min))
    if filters.date_to:
        stmt = stmt.where(AuditLog.created_at <= datetime.combine(filters.date_to, time.max))
    if filters.audit_action:
        stmt = stmt.where(AuditLog.action == filters.audit_action)
    total = total_for(db, stmt)
    logs = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    rows = [
        {
            "id": item.id,
            "action": item.action,
            "user": item.actor.full_name_ar if item.actor else "",
            "entity_type": item.entity_type,
            "entity_id": item.entity_id,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "ip_address": item.ip_address,
            "result": "ناجح",
            "old_value": (item.metadata_json or {}).get("old_value"),
            "new_value": (item.metadata_json or {}).get("new_value"),
        }
        for item in logs
    ]
    log_report_action(db, actor, "report_viewed", "audit", filters.compact(), request)
    db.commit()
    return {"summary": {"total_logs": total}, "items": rows, "charts": {"by_action": audit_action_counts(db, filters)}, "pagination": {"page": page, "page_size": page_size, "total": total}}


@router.get("/saved")
def list_saved_reports(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    stmt = select(SavedReport).order_by(SavedReport.is_favorite.desc(), SavedReport.updated_at.desc())
    if actor.role != UserRole.SUPER_ADMIN:
        stmt = stmt.where(SavedReport.created_by == actor.id)
    return [serialize_saved_report(item) for item in db.scalars(stmt).all()]


@router.post("/saved")
def create_saved_report(payload: SavedReportPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = SavedReport(**payload.model_dump(), created_by=actor.id)
    db.add(item)
    db.flush()
    log_report_action(db, actor, "saved_report_created", payload.report_type, payload.filters_json, request)
    db.commit()
    db.refresh(item)
    return serialize_saved_report(item)


@router.put("/saved/{report_id}")
def update_saved_report(report_id: int, payload: SavedReportPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(SavedReport, report_id)
    if not item or (actor.role != UserRole.SUPER_ADMIN and item.created_by != actor.id):
        raise HTTPException(status_code=404, detail="Report not found")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    log_report_action(db, actor, "saved_report_updated", payload.report_type, payload.filters_json, request)
    db.commit()
    return serialize_saved_report(item)


@router.delete("/saved/{report_id}")
def delete_saved_report(report_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(SavedReport, report_id)
    if not item or (actor.role != UserRole.SUPER_ADMIN and item.created_by != actor.id):
        raise HTTPException(status_code=404, detail="Report not found")
    report_type = item.report_type
    db.delete(item)
    log_report_action(db, actor, "saved_report_deleted", report_type, None, request)
    db.commit()
    return {"ok": True}


@router.post("/saved/{report_id}/run")
def run_saved_report(report_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(SavedReport, report_id)
    if not item or (actor.role != UserRole.SUPER_ADMIN and item.created_by != actor.id):
        raise HTTPException(status_code=404, detail="Report not found")
    filters = ReportFilters(**(item.filters_json or {}))
    return run_report_by_type(item.report_type, db, actor, filters, request)


@router.get("/templates")
def list_report_templates(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    ensure_default_templates(db, actor)
    db.commit()
    return [serialize_template(item) for item in db.scalars(select(ReportTemplate).order_by(ReportTemplate.name_ar)).all()]


@router.post("/templates")
def create_report_template(payload: ReportTemplatePayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    ensure_sensitive_report_allowed(actor, "users-permissions")
    item = ReportTemplate(**payload.model_dump(), created_by=actor.id)
    db.add(item)
    db.flush()
    log_report_action(db, actor, "report_template_created", payload.report_type, payload.default_filters_json, request)
    db.commit()
    db.refresh(item)
    return serialize_template(item)


@router.put("/templates/{template_id}")
def update_report_template(template_id: int, payload: ReportTemplatePayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    ensure_sensitive_report_allowed(actor, "users-permissions")
    item = db.get(ReportTemplate, template_id)
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    log_report_action(db, actor, "report_template_updated", payload.report_type, payload.default_filters_json, request)
    db.commit()
    return serialize_template(item)


@router.delete("/templates/{template_id}")
def delete_report_template(template_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    ensure_sensitive_report_allowed(actor, "users-permissions")
    item = db.get(ReportTemplate, template_id)
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    item.is_active = False
    log_report_action(db, actor, "report_template_disabled", item.report_type, None, request)
    db.commit()
    return {"ok": True}


@router.post("/templates/{template_id}/run")
def run_report_template(template_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(ReportTemplate, template_id)
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    filters = ReportFilters(**(item.default_filters_json or {}))
    return run_report_by_type(item.report_type, db, actor, filters, request)


@router.get("/scheduled")
def list_scheduled_reports(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    stmt = select(ScheduledReport).options(selectinload(ScheduledReport.template)).order_by(ScheduledReport.updated_at.desc())
    if actor.role != UserRole.SUPER_ADMIN:
        stmt = stmt.where(ScheduledReport.created_by == actor.id)
    return [serialize_schedule(item) for item in db.scalars(stmt).all()]


@router.post("/scheduled")
def create_scheduled_report(payload: ScheduledReportPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = ScheduledReport(**payload.model_dump(), created_by=actor.id)
    db.add(item)
    db.flush()
    log_report_action(db, actor, "scheduled_report_created", "scheduled", payload.model_dump(), request)
    db.commit()
    db.refresh(item)
    return serialize_schedule(item)


@router.put("/scheduled/{scheduled_id}")
def update_scheduled_report(scheduled_id: int, payload: ScheduledReportPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(ScheduledReport, scheduled_id)
    if not item or (actor.role != UserRole.SUPER_ADMIN and item.created_by != actor.id):
        raise HTTPException(status_code=404, detail="Schedule not found")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    log_report_action(db, actor, "scheduled_report_updated", "scheduled", payload.model_dump(), request)
    db.commit()
    return serialize_schedule(item)


@router.delete("/scheduled/{scheduled_id}")
def delete_scheduled_report(scheduled_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    item = db.get(ScheduledReport, scheduled_id)
    if not item or (actor.role != UserRole.SUPER_ADMIN and item.created_by != actor.id):
        raise HTTPException(status_code=404, detail="Schedule not found")
    db.delete(item)
    log_report_action(db, actor, "scheduled_report_deleted", "scheduled", None, request)
    db.commit()
    return {"ok": True}


@router.get("/export/excel")
def export_excel_center(
    request: Request,
    report_type: str = Query(default="requests"),
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
):
    return export_report_response("excel", report_type, db, actor, filters, request)


@router.get("/export/pdf")
def export_pdf_center(
    request: Request,
    report_type: str = Query(default="requests"),
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
):
    return export_report_response("pdf", report_type, db, actor, filters, request)


@router.get("/requests.xlsx")
def export_excel_legacy(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
):
    return export_report_response("excel", "requests", db, actor, filters, request)


@router.get("/requests.pdf")
def export_pdf_legacy(
    request: Request,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
    filters: ReportFilters = Depends(get_report_filters),
):
    return export_report_response("pdf", "requests", db, actor, filters, request)


def export_report_response(export_format: str, report_type: str, db: Session, actor: User, filters: ReportFilters, request: Request):
    ensure_sensitive_report_allowed(actor, report_type)
    if report_type != "requests":
        # Phase 1 keeps Excel/PDF export stable by exporting the filtered request dataset used by the dashboard.
        report_type = "requests"
    stmt = request_statement(db, actor, filters, with_options=True).limit(5000)
    items = db.scalars(stmt).all()
    if export_format == "excel":
        stream = build_excel_report(items, db, actor, filters)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "qib-requests-report.xlsx"
        action = "report_export_excel"
    else:
        stream = build_pdf_report(items, db, actor, filters)
        media_type = "application/pdf"
        filename = "qib-requests-report.pdf"
        action = "report_export_pdf"
    db.add(
        ReportExportLog(
            report_type=report_type,
            export_format=export_format,
            filters_json=filters.compact(),
            exported_by=actor.id,
            ip_address=request.client.host if request.client else None,
        )
    )
    log_report_action(db, actor, action, report_type, filters.compact(), request)
    db.commit()
    return StreamingResponse(stream, media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


def report_rows(items: Iterable[ServiceRequest], type_map: dict[int, RequestTypeSetting] | None = None) -> list[list[str]]:
    type_map = type_map or {}
    rows = []
    for item in items:
        rows.append(
            [
                str(getattr(item, "request_number", "") or ""),
                str(getattr(item, "title", "") or ""),
                str(requester_name(item) or ""),
                str(department_name(item) or ""),
                str(request_type_label(item, type_map) or ""),
                str(specialized_section_label(item, type_map) or ""),
                str(assigned_name(item) or ""),
                label(getattr(item, "status", None), STATUS_LABELS),
                label(getattr(item, "priority", None), PRIORITY_LABELS),
                datetime_label(getattr(item, "created_at", None)),
                datetime_label(getattr(item, "closed_at", None)),
                sla_label(sla_status(item)),
            ]
        )
    return rows


def build_excel_report(
    items: Iterable[ServiceRequest],
    db: Session | None = None,
    actor: User | None = None,
    filters: ReportFilters | None = None,
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    items = list(items)
    type_map = request_type_map(db, [getattr(item, "request_type_id", None) for item in items]) if db else {}
    actor_label = (getattr(actor, "full_name_ar", None) or getattr(actor, "email", None) or "النظام")
    filters_label = filters.compact() if filters else {}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "مركز التقارير"
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A5"
    sheet.append(["مركز التقارير - تقرير الطلبات"])
    sheet.append(["تم الإنشاء بواسطة", actor_label, "تاريخ الإنشاء", datetime.now(timezone.utc).strftime("%Y/%m/%d %H:%M")])
    sheet.append(["الفلاتر", str(filters_label or "بدون فلاتر")])
    sheet.append(EXCEL_HEADERS)
    header_row = 4
    header_alignment = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
    body_alignment = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=14)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6337")
        cell.alignment = header_alignment
    for row in report_rows(items, type_map):
        sheet.append(row)
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = body_alignment
    widths = [18, 34, 26, 26, 24, 24, 24, 18, 14, 22, 22, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_pdf_report(
    items: Iterable[ServiceRequest],
    db: Session | None = None,
    actor: User | None = None,
    filters: ReportFilters | None = None,
) -> BytesIO:
    items = list(items)
    stream = BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    font_name = register_arabic_pdf_font()
    general = db.scalar(select(SettingsGeneral).limit(1)) if db else None
    theme = pdf_theme(general)
    pdf.setTitle("مركز التقارير - تقرير الطلبات")
    left = 36
    right = A4[0] - 36
    actor_name = (getattr(actor, "full_name_ar", None) or getattr(actor, "email", None) or "النظام")
    y = draw_cover_header(
        pdf,
        theme,
        font_name,
        "مركز التقارير - تقرير الطلبات",
        f"تاريخ الطباعة: {format_pdf_datetime(datetime.now(timezone.utc), theme.timezone)}",
    )
    y = draw_section_header(pdf, theme, font_name, "ملخص التقرير", left, right, y)
    pdf.setFillColorRGB(0.98, 0.99, 1)
    pdf.setStrokeColorRGB(0.88, 0.91, 0.94)
    pdf.roundRect(left, y - 54, right - left, 54, 5, fill=1, stroke=1)
    draw_text(pdf, font_name, "عدد الطلبات", right - 14, y - 18, 8, (0.45, 0.5, 0.58))
    draw_ltr_text(pdf, font_name, str(len(items)), right - 120, y - 36, 13, (0.06, 0.09, 0.16))
    draw_text(pdf, font_name, "تم إنشاء التقرير بواسطة", left + 230, y - 18, 8, (0.45, 0.5, 0.58))
    draw_text(pdf, font_name, actor_name, left + 230, y - 36, 10, (0.06, 0.09, 0.16))
    y -= 78
    y = draw_section_header(pdf, theme, font_name, "قائمة الطلبات", left, right, y)
    columns = [
        ("رقم الطلب", right - 8, 92),
        ("العنوان", right - 112, 150),
        ("الموظف", right - 274, 116),
        ("الحالة", right - 402, 82),
        ("SLA", left + 92, 86),
    ]

    def draw_table_header(current_y: float) -> float:
        pdf.setFillColorRGB(0.96, 0.98, 1)
        pdf.setStrokeColorRGB(0.88, 0.91, 0.94)
        pdf.roundRect(left, current_y - 28, right - left, 28, 5, fill=1, stroke=1)
        for header, x, _ in columns:
            draw_text(pdf, font_name, header, x, current_y - 18, 8, (0.45, 0.5, 0.58))
        return current_y - 34

    y = draw_table_header(y)
    for item in items[:120]:
        if y < 58:
            draw_footer(pdf, font_name, f"طُبع بواسطة: {actor_name}", left, right)
            pdf.showPage()
            y = draw_page_header(pdf, theme, font_name, "مركز التقارير")
            y = draw_table_header(y)
        pdf.setStrokeColorRGB(0.92, 0.94, 0.96)
        pdf.line(left, y - 7, right, y - 7)
        draw_ltr_text(pdf, font_name, str(getattr(item, "request_number", "") or "")[:18], right - 95, y + 5, 9, (0.06, 0.09, 0.16))
        draw_text(pdf, font_name, str(getattr(item, "title", "") or "-")[:34], right - 112, y + 5, 9, (0.06, 0.09, 0.16))
        draw_text(pdf, font_name, requester_name(item)[:24], right - 274, y + 5, 9, (0.06, 0.09, 0.16))
        draw_text(pdf, font_name, label(getattr(item, "status", None), STATUS_LABELS), right - 402, y + 5, 9, theme.brand_dark)
        draw_text(pdf, font_name, sla_label(sla_status(item)), left + 92, y + 5, 8, (0.36, 0.42, 0.48))
        y -= 24
    draw_footer(pdf, font_name, f"طُبع بواسطة: {actor_name}", left, right)
    pdf.save()
    stream.seek(0)
    return stream


def messaging_count(db: Session, actor: User, filters: ReportFilters) -> int:
    return total_for(db, apply_message_scope(apply_message_filters(select(InternalMessage), filters), actor).distinct())


def message_count_by_type(db: Session, actor: User, filters: ReportFilters, message_type: str) -> int:
    return total_for(db, apply_message_scope(apply_message_filters(select(InternalMessage).where(InternalMessage.message_type == message_type), filters), actor).distinct())


def unread_message_count(db: Session, actor: User, filters: ReportFilters) -> int:
    stmt = (
        select(InternalMessage)
        .join(InternalMessageRecipient, InternalMessageRecipient.message_id == InternalMessage.id)
        .where(InternalMessageRecipient.is_read == False)
    )
    stmt = apply_message_filters(stmt, filters)
    if actor.role not in {UserRole.SUPER_ADMIN, UserRole.EXECUTIVE}:
        stmt = stmt.where(InternalMessageRecipient.recipient_id == actor.id)
    return total_for(db, stmt.distinct())


def message_group_counts(db: Session, actor: User, filters: ReportFilters, column) -> list[dict]:
    stmt = select(column, func.count(InternalMessage.id))
    stmt = apply_message_filters(stmt, filters)
    stmt = apply_message_scope(stmt, actor).group_by(column)
    return chart_from_counts(db.execute(stmt).all())


def user_map(db: Session, ids: Iterable[int | None]) -> dict[int, str]:
    clean_ids = {item for item in ids if item}
    if not clean_ids:
        return {}
    return {item.id: item.full_name_ar for item in db.scalars(select(User).where(User.id.in_(clean_ids))).all()}


def user_role_counts(db: Session) -> list[dict]:
    rows = db.execute(select(User.role, func.count(User.id)).group_by(User.role)).all()
    return chart_from_counts([(label(role, {}), count) for role, count in rows])


def user_department_counts(db: Session) -> list[dict]:
    rows = db.execute(select(Department.name_ar, func.count(User.id)).join(User, User.department_id == Department.id, isouter=True).group_by(Department.name_ar)).all()
    return chart_from_counts(rows)


def requests_by_month(db: Session, actor: User, filters: ReportFilters) -> list[dict]:
    stmt = request_statement(db, actor, filters, ordered=False)
    items = db.scalars(stmt.limit(2000)).all()
    counts: dict[str, int] = {}
    for item in items:
        key = item.created_at.strftime("%Y-%m") if item.created_at else "غير محدد"
        counts[key] = counts.get(key, 0) + 1
    return [{"label": key, "count": counts[key]} for key in sorted(counts)]


def sla_trend(db: Session, actor: User, filters: ReportFilters) -> list[dict]:
    stmt = request_statement(db, actor, filters, ordered=False)
    items = db.scalars(stmt.limit(2000)).all()
    buckets: dict[str, dict[str, int]] = {}
    for item in items:
        key = item.created_at.strftime("%Y-%m") if item.created_at else "غير محدد"
        bucket = buckets.setdefault(key, {"met": 0, "breached": 0})
        bucket["breached" if sla_status(item) == "breached" else "met"] += 1
    return [{"label": key, **buckets[key]} for key in sorted(buckets)]


def sla_status_counts(items: Iterable[ServiceRequest]) -> list[dict]:
    counts: dict[str, int] = {}
    for item in items:
        value = sla_status(item)
        counts[sla_label(value)] = counts.get(sla_label(value), 0) + 1
    return [{"label": key, "count": count} for key, count in counts.items()]


def breached_today(db: Session, actor: User, filters: ReportFilters) -> int:
    today = date.today()
    scoped = request_statement(db, actor, filters, ordered=False)
    return total_for(
        db,
        scoped.where(
            ServiceRequest.sla_due_at >= datetime.combine(today, time.min),
            ServiceRequest.sla_due_at <= datetime.combine(today, time.max),
            ServiceRequest.sla_due_at < datetime.now(timezone.utc),
            ServiceRequest.status.not_in(FINAL_REQUEST_STATUSES),
        ),
    )


def approval_status_counts(db: Session, request_ids) -> list[dict]:
    rows = db.execute(
        select(RequestApprovalStep.status, func.count(RequestApprovalStep.id))
        .where(RequestApprovalStep.request_id.in_(select(request_ids.c.id)))
        .group_by(RequestApprovalStep.status)
    ).all()
    return chart_from_counts(rows)


def audit_action_counts(db: Session, filters: ReportFilters) -> list[dict]:
    stmt = select(AuditLog.action, func.count(AuditLog.id)).group_by(AuditLog.action)
    if filters.date_from:
        stmt = stmt.where(AuditLog.created_at >= datetime.combine(filters.date_from, time.min))
    if filters.date_to:
        stmt = stmt.where(AuditLog.created_at <= datetime.combine(filters.date_to, time.max))
    return chart_from_counts(db.execute(stmt).all())


def run_report_by_type(report_type: str, db: Session, actor: User, filters: ReportFilters, request: Request) -> dict:
    if report_type == "approvals":
        return approval_reports(request, db, actor, filters, page=1, page_size=25)
    if report_type == "sla":
        return sla_reports(request, db, actor, filters, page=1, page_size=25)
    if report_type == "users-permissions":
        return users_permissions_reports(request, db, actor, filters, page=1, page_size=25)
    if report_type == "messaging":
        return messaging_reports(request, db, actor, filters, page=1, page_size=25)
    if report_type == "attachments":
        return attachment_reports(request, db, actor, filters, page=1, page_size=25)
    if report_type == "audit":
        return audit_reports(request, db, actor, filters, page=1, page_size=25)
    return request_reports(request, db, actor, filters, page=1, page_size=25)
