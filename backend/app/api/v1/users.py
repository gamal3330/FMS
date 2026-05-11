from datetime import datetime, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ValidationError
from sqlalchemy import and_, func, or_
from sqlalchemy import delete, select
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, require_roles
from app.core.security import get_password_hash, verify_password
from app.db.session import get_db
from app.models.audit import AuditLog
from app.models.ai import AIFeaturePermission
from app.models.enums import UserRole
from app.models.request import RequestApprovalStep
from app.models.settings import PortalSetting, SecurityPolicy, SettingsGeneral, SpecializedSection
from app.models.settings import WorkflowTemplateStep
from app.models.user import (
    AccessReview,
    AccessReviewItem,
    ActionPermission,
    Department,
    Role,
    ScreenPermission,
    User,
    UserDelegation,
    UserImportBatch,
    UserImportError,
    UserLoginAttempt,
    UserSession,
)
from app.schemas.settings import SettingsDepartmentCreate
from app.schemas.user import DepartmentRead, PasswordReset, UserCreate, UserRead, UserUpdate
from app.services.audit import write_audit

router = APIRouter(prefix="/users", tags=["Users"])


SCREEN_DEFINITIONS = [
    {"key": "dashboard", "label": "إحصائيات"},
    {"key": "requests", "label": "الطلبات"},
    {"key": "approvals", "label": "الموافقات"},
    {"key": "messages", "label": "المراسلات الداخلية"},
    {"key": "documents", "label": "مكتبة الوثائق"},
    {"key": "reports", "label": "التقارير"},
    {"key": "request_types", "label": "إدارة أنواع الطلبات"},
    {"key": "users", "label": "المستخدمون والصلاحيات"},
    {"key": "departments", "label": "الإدارات"},
    {"key": "specialized_sections", "label": "الأقسام المختصة"},
    {"key": "messaging_settings", "label": "إعدادات المراسلات"},
    {"key": "ai_settings", "label": "إعدادات الذكاء الاصطناعي"},
    {"key": "database_settings", "label": "إعدادات قاعدة البيانات"},
    {"key": "health_monitoring", "label": "مراقبة صحة النظام"},
    {"key": "document_settings", "label": "إعدادات الوثائق"},
    {"key": "settings", "label": "الإعدادات"},
]

ALL_SCREEN_KEYS = {item["key"] for item in SCREEN_DEFINITIONS}
MANAGEMENT_SCREEN_KEYS = {"dashboard", "requests", "approvals", "messages", "documents", "reports", "request_types", "users", "departments", "specialized_sections", "messaging_settings", "ai_settings", "database_settings", "health_monitoring", "document_settings", "settings"}
EMPLOYEE_SCREEN_KEYS = {"requests", "approvals", "messages", "documents", "reports"}
DASHBOARD_SCREEN_ROLES = {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER, UserRole.EXECUTIVE}
ROLE_LABELS = {
    UserRole.EMPLOYEE.value: "موظف",
    UserRole.DIRECT_MANAGER.value: "مدير مباشر",
    UserRole.IT_STAFF.value: "مختص تنفيذ",
    UserRole.DEPARTMENT_MANAGER.value: "مدير إدارة",
    UserRole.INFOSEC.value: "أمن المعلومات (دور قديم)",
    UserRole.EXECUTIVE.value: "الإدارة التنفيذية",
    UserRole.SUPER_ADMIN.value: "مدير النظام",
}

ACTIVE_USER_ROLES = [
    UserRole.EMPLOYEE,
    UserRole.DIRECT_MANAGER,
    UserRole.IT_STAFF,
    UserRole.DEPARTMENT_MANAGER,
    UserRole.EXECUTIVE,
    UserRole.SUPER_ADMIN,
]
HIDDEN_LEGACY_ROLE_CODES = {"information_security", "it_manager", "department_manager"}
HIDDEN_LEGACY_ROLE_PREFIXES = (
    "information_security_copy",
    "information_security_legacy",
    "it_manager_copy",
    "it_manager_legacy",
    "department_manager_copy",
    "department_manager_legacy",
)


def is_hidden_legacy_role(role: Role | None) -> bool:
    if not role:
        return False
    code = str(role.code or role.name or "")
    return code in HIDDEN_LEGACY_ROLE_CODES or any(code.startswith(prefix) for prefix in HIDDEN_LEGACY_ROLE_PREFIXES)


def is_hidden_legacy_role_code(code: str | None) -> bool:
    value = str(code or "")
    return value in HIDDEN_LEGACY_ROLE_CODES or any(value.startswith(prefix) for prefix in HIDDEN_LEGACY_ROLE_PREFIXES)

PERMISSION_LEVELS = ["no_access", "view", "create", "edit", "delete", "export", "manage"]
ACTION_DEFINITIONS = [
    {"code": "approve_request", "label": "اعتماد الطلب", "group": "الموافقات", "dangerous": False},
    {"code": "reject_request", "label": "رفض الطلب", "group": "الموافقات", "dangerous": False},
    {"code": "return_request_for_edit", "label": "إرجاع الطلب للتعديل", "group": "الموافقات", "dangerous": False},
    {"code": "execute_request", "label": "تنفيذ الطلب", "group": "الطلبات", "dangerous": False},
    {"code": "close_request", "label": "إغلاق الطلب", "group": "الطلبات", "dangerous": False},
    {"code": "reopen_request", "label": "إعادة فتح الطلب", "group": "الطلبات", "dangerous": True},
    {"code": "print_request_pdf", "label": "طباعة PDF", "group": "التقارير", "dangerous": False},
    {"code": "export_reports", "label": "تصدير التقارير", "group": "التقارير", "dangerous": False},
    {"code": "manage_request_types", "label": "إدارة أنواع الطلبات", "group": "الإعدادات", "dangerous": True},
    {"code": "manage_workflows", "label": "إدارة مسارات الموافقة", "group": "الإعدادات", "dangerous": True},
    {"code": "manage_database", "label": "إدارة قاعدة البيانات", "group": "قاعدة البيانات", "dangerous": True},
    {"code": "restore_database", "label": "استعادة قاعدة البيانات", "group": "قاعدة البيانات", "dangerous": True},
    {"code": "reset_database", "label": "إعادة ضبط قاعدة البيانات", "group": "قاعدة البيانات", "dangerous": True},
    {"code": "manage_ai_settings", "label": "إدارة الذكاء الاصطناعي", "group": "الذكاء الاصطناعي", "dangerous": True},
    {"code": "audit_messages", "label": "تدقيق المراسلات", "group": "المراسلات", "dangerous": True},
    {"code": "view_confidential_messages", "label": "عرض الرسائل السرية", "group": "المراسلات", "dangerous": True},
]


class ScreenPermissionsPayload(BaseModel):
    screens: list[str]


class ScreenPermissionsRead(BaseModel):
    screens: list[str]
    available_screens: list[dict[str, str]]


class PermissionLevelPayload(BaseModel):
    permissions: dict[str, str]


class ActionPermissionPayload(BaseModel):
    permissions: dict[str, bool]
    confirmation_text: str | None = None


class RolePayload(BaseModel):
    name_ar: str
    name_en: str | None = None
    code: str
    description: str | None = None
    is_active: bool = True


class PasswordConfirmPayload(BaseModel):
    admin_password: str | None = None
    confirmation_text: str | None = None


class DelegationPayload(BaseModel):
    delegator_user_id: int
    delegate_user_id: int
    delegation_scope: str = "approvals_only"
    start_date: datetime
    end_date: datetime
    is_active: bool = True
    reason: str | None = None


class BulkAssignDepartmentPayload(BaseModel):
    user_ids: list[int]
    department_id: int


class BulkAssignManagerPayload(BaseModel):
    user_ids: list[int]
    manager_id: int


class ImportConfirmPayload(BaseModel):
    batch_id: int
    import_valid_only: bool = True
    confirmation_text: str | None = None


IMPORT_COLUMNS = [
    ("employee_id", "الرقم الوظيفي", True),
    ("username", "اسم المستخدم", False),
    ("full_name_ar", "الاسم العربي", True),
    ("full_name_en", "الاسم الإنجليزي", True),
    ("email", "البريد الإلكتروني", True),
    ("mobile", "رقم الجوال", False),
    ("role", "الصلاحية", True),
    ("department_code", "كود الإدارة", True),
    ("manager_employee_id", "الرقم الوظيفي للمدير المباشر", False),
    ("administrative_section", "كود القسم المختص", False),
    ("password", "كلمة المرور المؤقتة", False),
    ("is_active", "حساب نشط", False),
]
IMPORT_HEADER_ALIASES = {key: key for key, _, _ in IMPORT_COLUMNS} | {label: key for key, label, _ in IMPORT_COLUMNS}
DEFAULT_IMPORTED_PASSWORD = "Change@12345"


def temporary_password_from_policy(policy: SecurityPolicy) -> str:
    return policy.temporary_password or DEFAULT_IMPORTED_PASSWORD


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def parse_bool(value, default: bool = True) -> bool:
    text = cell_text(value).lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "active", "نشط", "نعم"}:
        return True
    if text in {"0", "false", "no", "n", "inactive", "disabled", "معطل", "لا"}:
        return False
    raise ValueError("قيمة حالة الحساب يجب أن تكون نعم/لا أو true/false")


def import_error(row: int, field: str, message: str) -> dict[str, str | int]:
    return {"row": row, "field": field, "message": message}


def default_screens_for_role(role: UserRole) -> list[str]:
    if role in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER}:
        return sorted(MANAGEMENT_SCREEN_KEYS)
    if role == UserRole.EXECUTIVE:
        return ["dashboard", "requests", "approvals", "messages", "reports"]
    if role in {UserRole.INFOSEC, UserRole.IT_STAFF}:
        return ["requests", "approvals", "messages", "reports"]
    return sorted(EMPLOYEE_SCREEN_KEYS)


def available_screens_for_user(user: User) -> list[dict[str, str]]:
    if user.role in DASHBOARD_SCREEN_ROLES:
        return SCREEN_DEFINITIONS
    return [screen for screen in SCREEN_DEFINITIONS if screen["key"] != "dashboard"]


def get_screen_permission_setting(db: Session, user_id: int) -> PortalSetting | None:
    return db.scalar(select(PortalSetting).where(PortalSetting.category == "screen_permissions", PortalSetting.setting_key == str(user_id)))


def read_user_screens(db: Session, user: User) -> list[str]:
    setting = get_screen_permission_setting(db, user.id)
    if not setting:
        return default_screens_for_role(user.role)
    value = setting.setting_value if isinstance(setting.setting_value, dict) else {}
    screens = value.get("screens", [])
    clean_screens = [screen for screen in screens if screen in ALL_SCREEN_KEYS]
    if user.role not in DASHBOARD_SCREEN_ROLES:
        clean_screens = [screen for screen in clean_screens if screen != "dashboard"]
    if "messages_permission_initialized" not in value and "messages" in default_screens_for_role(user.role) and "messages" not in clean_screens:
        clean_screens.append("messages")
    if "documents_permission_initialized" not in value and "documents" in default_screens_for_role(user.role) and "documents" not in clean_screens:
        clean_screens.append("documents")
    if user.role in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER} and "settings" in clean_screens and "health_monitoring" not in clean_screens:
        clean_screens.append("health_monitoring")
    if user.role in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER} and "settings" in clean_screens and "document_settings" not in clean_screens:
        clean_screens.append("document_settings")
    return clean_screens


def can_view_users_screen(db: Session, user: User) -> bool:
    return user.role == UserRole.DIRECT_MANAGER or permission_level_allows(effective_screen_permission_level(db, user, "users"), "view")


def require_users_screen_view(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> User:
    if not can_view_users_screen(db, current_user):
        raise HTTPException(status_code=403, detail="لا تملك صلاحية عرض شاشة المستخدمين والصلاحيات")
    return current_user


def has_active_approval_delegation(db: Session, user: User) -> bool:
    now = datetime.now(timezone.utc)
    return (
        db.scalar(
            select(UserDelegation.id)
            .where(
                UserDelegation.delegate_user_id == user.id,
                UserDelegation.is_active == True,
                UserDelegation.delegation_scope.in_(["approvals_only", "all_allowed_actions"]),
                UserDelegation.start_date <= now,
                UserDelegation.end_date >= now,
            )
            .limit(1)
        )
        is not None
    )


def save_user_screens(db: Session, user: User, screens: list[str], actor: User) -> None:
    clean_screens = sorted({screen for screen in screens if screen in ALL_SCREEN_KEYS})
    if user.role not in DASHBOARD_SCREEN_ROLES:
        clean_screens = [screen for screen in clean_screens if screen != "dashboard"]
    setting = get_screen_permission_setting(db, user.id)
    if not setting:
        setting = PortalSetting(category="screen_permissions", setting_key=str(user.id), setting_value={})
        db.add(setting)
    setting.setting_value = {"screens": clean_screens, "messages_permission_initialized": True, "documents_permission_initialized": True}
    setting.updated_by_id = actor.id


def validate_user_links(db: Session, payload: UserCreate | UserUpdate, user_id: int | None = None) -> None:
    if payload.department_id and not db.get(Department, payload.department_id):
        raise HTTPException(status_code=400, detail="Department not found")
    if payload.manager_id:
        if user_id and payload.manager_id == user_id:
            raise HTTPException(status_code=400, detail="User cannot be their own direct manager")
        manager = db.get(User, payload.manager_id)
        if not manager or not manager.is_active:
            raise HTTPException(status_code=400, detail="Direct manager not found or inactive")
        if manager.role not in {UserRole.DIRECT_MANAGER, UserRole.DEPARTMENT_MANAGER, UserRole.SUPER_ADMIN, UserRole.EXECUTIVE}:
            raise HTTPException(status_code=400, detail="Direct manager must have a manager-level role")
        if payload.department_id and manager.role == UserRole.DIRECT_MANAGER and manager.department_id != payload.department_id:
            raise HTTPException(status_code=400, detail="Direct manager must belong to the same department")
    if payload.role == UserRole.IT_STAFF and not payload.administrative_section:
        raise HTTPException(status_code=400, detail="القسم المختص مطلوب لمختص التنفيذ")
    if payload.administrative_section:
        section = db.scalar(select(SpecializedSection).where(SpecializedSection.code == payload.administrative_section, SpecializedSection.is_active == True))
        if not section:
            raise HTTPException(status_code=400, detail="القسم المختص غير موجود أو غير نشط")


def ensure_role_assignment_allowed(actor: User, role: UserRole) -> None:
    if role not in ACTIVE_USER_ROLES:
        raise HTTPException(status_code=422, detail="هذا الدور قديم وغير متاح للاستخدام. استخدم الأدوار المعتمدة الحالية.")
    if actor.role != UserRole.SUPER_ADMIN and role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can assign Super Admin role")


def ensure_user_unique(db: Session, payload: UserCreate | UserUpdate, user_id: int | None = None) -> None:
    conditions = [User.employee_id == payload.employee_id, User.email == str(payload.email)]
    if payload.username:
        conditions.append(User.username == payload.username)
    stmt = select(User).where(or_(*conditions))
    if user_id:
        stmt = stmt.where(User.id != user_id)
    duplicate = db.scalar(stmt)
    if not duplicate:
        return
    if duplicate.employee_id == payload.employee_id:
        raise HTTPException(status_code=409, detail="الرقم الوظيفي مستخدم من قبل")
    if duplicate.email == str(payload.email):
        raise HTTPException(status_code=409, detail="البريد الإلكتروني مستخدم من قبل")
    if payload.username and duplicate.username == payload.username:
        raise HTTPException(status_code=409, detail="اسم المستخدم مستخدم من قبل")


def security_policy(db: Session) -> SecurityPolicy:
    item = db.scalar(select(SecurityPolicy).limit(1))
    if not item:
        item = SecurityPolicy()
        db.add(item)
        db.flush()
    return item


def validate_password_policy(password: str, policy: SecurityPolicy) -> None:
    if len(password) < policy.password_min_length:
        raise HTTPException(status_code=422, detail=f"كلمة المرور يجب ألا تقل عن {policy.password_min_length} أحرف")
    if policy.require_uppercase and not any(char.isupper() for char in password):
        raise HTTPException(status_code=422, detail="كلمة المرور يجب أن تحتوي على حرف كبير")
    if policy.require_numbers and not any(char.isdigit() for char in password):
        raise HTTPException(status_code=422, detail="كلمة المرور يجب أن تحتوي على رقم")
    if policy.require_special_chars and not any(not char.isalnum() for char in password):
        raise HTTPException(status_code=422, detail="كلمة المرور يجب أن تحتوي على رمز خاص")


def require_actor_password(actor: User, password: str | None) -> None:
    if not password or not verify_password(password, actor.hashed_password):
        raise HTTPException(status_code=403, detail="كلمة مرور المدير غير صحيحة")


def normalize_local_datetime(db: Session, value: datetime) -> datetime:
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc)
    general = db.scalar(select(SettingsGeneral).limit(1))
    timezone_name = general.timezone if general and general.timezone else "Asia/Qatar"
    try:
        local_tz = ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError:
        local_tz = ZoneInfo("Asia/Qatar")
    return value.replace(tzinfo=local_tz).astimezone(timezone.utc)


def client_ip(request: Request) -> str | None:
    return request.client.host if request.client else None


def request_user_agent(request: Request) -> str | None:
    value = request.headers.get("user-agent")
    return value[:255] if value else None


def role_label(role: str | UserRole) -> str:
    return ROLE_LABELS.get(str(role), str(role))


def role_to_dict(role: Role) -> dict:
    code = role.code or role.name
    return {
        "id": role.id,
        "role_name_ar": role.name_ar or role.label_ar,
        "role_name_en": role.name_en or role.name.replace("_", " ").title(),
        "code": code,
        "description": role.description,
        "is_system_role": role.is_system_role,
        "is_active": role.is_active,
        "users_count": 0,
        "created_at": role.created_at,
        "updated_at": role.updated_at,
    }


def user_display(user: User | None) -> dict | None:
    if not user:
        return None
    return {"id": user.id, "full_name_ar": user.full_name_ar, "email": user.email, "employee_id": user.employee_id}


def locked_now(user: User) -> bool:
    if getattr(user, "is_locked", False):
        return True
    if not user.locked_until:
        return False
    locked_until = user.locked_until.replace(tzinfo=timezone.utc) if user.locked_until.tzinfo is None else user.locked_until
    return locked_until > datetime.now(timezone.utc)


def ensure_not_last_super_admin(db: Session, user: User, next_role: UserRole | None = None, next_active: bool | None = None) -> None:
    is_losing_super_admin = user.role == UserRole.SUPER_ADMIN and (next_role and next_role != UserRole.SUPER_ADMIN or next_active is False)
    if not is_losing_super_admin:
        return
    count = db.scalar(select(func.count()).select_from(User).where(User.role == UserRole.SUPER_ADMIN, User.is_active == True, User.id != user.id)) or 0
    if count == 0:
        raise HTTPException(status_code=409, detail="لا يمكن تعطيل أو إزالة صلاحية آخر مدير نظام نشط")


def permission_level_to_flags(level: str) -> dict[str, bool]:
    return {
        "can_view": level in {"view", "create", "edit", "delete", "export", "manage"},
        "can_create": level in {"create", "edit", "delete", "manage"},
        "can_edit": level in {"edit", "delete", "manage"},
        "can_delete": level in {"delete", "manage"},
        "can_export": level in {"export", "manage"},
        "can_manage": level == "manage",
    }


def read_screen_permission_levels(db: Session, *, role_id: int | None = None, user_id: int | None = None, fallback_screens: list[str] | None = None) -> dict[str, str]:
    stmt = select(ScreenPermission)
    if role_id is not None:
        stmt = stmt.where(ScreenPermission.role_id == role_id, ScreenPermission.user_id.is_(None))
    elif user_id is not None:
        stmt = stmt.where(ScreenPermission.user_id == user_id)
    else:
        return {screen: "view" for screen in fallback_screens or []}
    rows = db.scalars(stmt).all()
    permissions = {
        row.screen_code: row.permission_level if row.permission_level in PERMISSION_LEVELS else "view"
        for row in rows
        if row.screen_code in ALL_SCREEN_KEYS
    }
    for screen in fallback_screens or []:
        permissions.setdefault(screen, "view")
    return permissions


def set_screen_permission_levels(db: Session, permissions: dict[str, str], *, role_id: int | None = None, user_id: int | None = None) -> None:
    for code, level in permissions.items():
        if code not in ALL_SCREEN_KEYS:
            continue
        clean_level = level if level in PERMISSION_LEVELS else "no_access"
        if role_id is not None:
            row = db.scalar(select(ScreenPermission).where(ScreenPermission.role_id == role_id, ScreenPermission.user_id.is_(None), ScreenPermission.screen_code == code))
            if not row:
                row = ScreenPermission(role_id=role_id, screen_code=code)
                db.add(row)
        elif user_id is not None:
            row = db.scalar(select(ScreenPermission).where(ScreenPermission.user_id == user_id, ScreenPermission.role_id.is_(None), ScreenPermission.screen_code == code))
            if not row:
                row = ScreenPermission(user_id=user_id, screen_code=code)
                db.add(row)
        else:
            continue
        row.permission_level = clean_level
        for field, value in permission_level_to_flags(clean_level).items():
            setattr(row, field, value)


def permission_level_allows(level: str | None, capability: str) -> bool:
    clean_level = level if level in PERMISSION_LEVELS else "no_access"
    return bool(permission_level_to_flags(clean_level).get(f"can_{capability}", False))


def role_record_for_user(db: Session, user: User) -> Role | None:
    if user.role_id:
        role = db.get(Role, user.role_id)
        if role:
            return role
    return db.scalar(select(Role).where(or_(Role.code == str(user.role), Role.name == str(user.role))))


def effective_screen_permission_level(db: Session, user: User, screen_code: str) -> str:
    if screen_code not in ALL_SCREEN_KEYS:
        return "no_access"
    if user.role in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER}:
        return "manage"

    user_permission = db.scalar(
        select(ScreenPermission).where(
            ScreenPermission.user_id == user.id,
            ScreenPermission.role_id.is_(None),
            ScreenPermission.screen_code == screen_code,
        )
    )
    if user_permission:
        return user_permission.permission_level if user_permission.permission_level in PERMISSION_LEVELS else "no_access"

    role = role_record_for_user(db, user)
    if role:
        role_permission = db.scalar(
            select(ScreenPermission).where(
                ScreenPermission.role_id == role.id,
                ScreenPermission.user_id.is_(None),
                ScreenPermission.screen_code == screen_code,
            )
        )
        if role_permission:
            return role_permission.permission_level if role_permission.permission_level in PERMISSION_LEVELS else "no_access"

    return "view" if screen_code in read_user_screens(db, user) else "no_access"


def require_users_screen_capability(db: Session, current_user: User, capability: str, detail: str) -> User:
    if not permission_level_allows(effective_screen_permission_level(db, current_user, "users"), capability):
        raise HTTPException(status_code=403, detail=detail)
    return current_user


def require_users_screen_create(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> User:
    return require_users_screen_capability(db, current_user, "create", "لا تملك صلاحية إضافة مستخدمين")


def require_users_screen_edit(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> User:
    return require_users_screen_capability(db, current_user, "edit", "لا تملك صلاحية تعديل المستخدمين")


def require_users_screen_manage(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> User:
    return require_users_screen_capability(db, current_user, "manage", "لا تملك صلاحية إدارة المستخدمين")


def role_by_code(db: Session, code: str) -> Role | None:
    return db.scalar(select(Role).where(or_(Role.code == code, Role.name == code)))


def role_screen_permission_setting(db: Session, role_code: str) -> PortalSetting | None:
    return db.scalar(select(PortalSetting).where(PortalSetting.category == "role_screen_permissions", PortalSetting.setting_key == role_code))


def read_role_screens(db: Session, role_code: str) -> list[str]:
    setting = role_screen_permission_setting(db, role_code)
    if setting and isinstance(setting.setting_value, dict):
        return [screen for screen in setting.setting_value.get("screens", []) if screen in ALL_SCREEN_KEYS]
    try:
        return default_screens_for_role(UserRole(role_code))
    except ValueError:
        return []


def save_role_screens(db: Session, role: Role, screens: list[str], actor: User) -> None:
    clean_screens = sorted({screen for screen in screens if screen in ALL_SCREEN_KEYS})
    setting = role_screen_permission_setting(db, role.code or role.name)
    if not setting:
        setting = PortalSetting(category="role_screen_permissions", setting_key=role.code or role.name, setting_value={})
        db.add(setting)
    setting.setting_value = {"screens": clean_screens}
    setting.updated_by_id = actor.id


def read_action_permissions(db: Session, *, role_id: int | None = None, user_id: int | None = None) -> dict[str, bool]:
    stmt = select(ActionPermission)
    if role_id:
        stmt = stmt.where(ActionPermission.role_id == role_id, ActionPermission.user_id.is_(None))
    elif user_id:
        stmt = stmt.where(ActionPermission.user_id == user_id)
    rows = db.scalars(stmt).all()
    return {row.action_code: row.is_allowed for row in rows}


def set_action_permissions(db: Session, permissions: dict[str, bool], actor: User, *, role_id: int | None = None, user_id: int | None = None) -> None:
    known = {item["code"] for item in ACTION_DEFINITIONS}
    for action_code, is_allowed in permissions.items():
        if action_code not in known:
            continue
        stmt = select(ActionPermission).where(ActionPermission.action_code == action_code)
        if role_id:
            stmt = stmt.where(ActionPermission.role_id == role_id, ActionPermission.user_id.is_(None))
        else:
            stmt = stmt.where(ActionPermission.user_id == user_id)
        row = db.scalar(stmt)
        if not row:
            row = ActionPermission(role_id=role_id, user_id=user_id, action_code=action_code)
            db.add(row)
        row.is_allowed = bool(is_allowed)
    write_audit(db, "action_permission_changed", "permission", actor=actor, metadata={"role_id": role_id, "user_id": user_id, "permissions": permissions})


@router.get("/import-template")
def download_users_import_template(db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    from openpyxl import Workbook

    default_password = temporary_password_from_policy(security_policy(db))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "users"
    sheet.append([label for _, label, _ in IMPORT_COLUMNS])
    sheet.append([
        "1001",
        "sample.user",
        "مستخدم تجريبي",
        "Sample User",
        "sample.user@example.com",
        "555000000",
        UserRole.EMPLOYEE.value,
        "IT",
        "",
        "",
        default_password,
        "نعم",
    ])

    notes = workbook.create_sheet("instructions")
    notes.append(["الحقل", "إلزامي", "ملاحظات"])
    notes.append(["الصلاحية", "نعم", "استخدم إحدى القيم من ورقة roles"])
    notes.append(["كود الإدارة", "نعم", "استخدم code من ورقة departments أو رقم id"])
    notes.append(["الرقم الوظيفي للمدير المباشر", "لا", "يمكن أن يشير إلى مستخدم موجود أو صف آخر داخل نفس الملف"])
    notes.append(["كود القسم المختص", "لمختص التنفيذ فقط", "مطلوب عند role = it_staff، استخدم ورقة specialized_sections"])
    notes.append(["كلمة المرور المؤقتة", "لا", f"إذا تركت فارغة سيتم استخدام كلمة المرور المؤقتة المعرفة في إعدادات الأمان ({default_password})"])
    notes.append(["حساب نشط", "لا", "القيم المقبولة: نعم/لا أو true/false أو 1/0"])

    roles_sheet = workbook.create_sheet("roles")
    roles_sheet.append(["role"])
    for role in ACTIVE_USER_ROLES:
        roles_sheet.append([role.value])

    departments_sheet = workbook.create_sheet("departments")
    departments_sheet.append(["id", "code", "name_ar", "name_en"])
    for department in db.scalars(select(Department).order_by(Department.name_ar)).all():
        departments_sheet.append([department.id, department.code or "", department.name_ar, department.name_en])

    sections_sheet = workbook.create_sheet("specialized_sections")
    sections_sheet.append(["code", "name_ar", "name_en"])
    for section in db.scalars(select(SpecializedSection).where(SpecializedSection.is_active == True).order_by(SpecializedSection.name_ar)).all():
        sections_sheet.append([section.code, section.name_ar, section.name_en])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users-import-template.xlsx"'},
    )


@router.post("/import")
async def import_users_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="يرجى رفع ملف Excel بصيغة .xlsx")

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="تعذر قراءة ملف Excel") from exc

    sheet = workbook["users"] if "users" in workbook.sheetnames else workbook.active
    headers = [IMPORT_HEADER_ALIASES.get(cell_text(value), "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    missing_headers = [label for key, label, required in IMPORT_COLUMNS if required and key not in headers]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"الأعمدة الإلزامية غير موجودة: {', '.join(missing_headers)}")

    rows: list[dict] = []
    errors: list[dict] = []
    departments = db.scalars(select(Department)).all()
    department_by_key = {}
    for department in departments:
        for key in [department.id, department.code, department.name_ar, department.name_en]:
            if key is not None and cell_text(key):
                department_by_key[cell_text(key).lower()] = department
    active_sections = {section.code for section in db.scalars(select(SpecializedSection).where(SpecializedSection.is_active == True)).all()}
    existing_users = db.scalars(select(User)).all()
    existing_by_employee = {user.employee_id: user for user in existing_users}
    existing_emails = {user.email.lower() for user in existing_users}
    existing_usernames = {user.username.lower() for user in existing_users if user.username}
    seen_employee_ids: set[str] = set()
    seen_emails: set[str] = set()
    seen_usernames: set[str] = set()
    policy = security_policy(db)
    default_password = temporary_password_from_policy(policy)

    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell_text(value) for value in values):
            continue
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]}
        employee_id = cell_text(row.get("employee_id"))
        username = cell_text(row.get("username")) or None
        email = cell_text(row.get("email")).lower()
        role_text = cell_text(row.get("role"))
        department_key = cell_text(row.get("department_code")).lower()
        password = cell_text(row.get("password")) or default_password
        manager_employee_id = cell_text(row.get("manager_employee_id")) or None
        administrative_section = cell_text(row.get("administrative_section")) or None

        for key, label, required in IMPORT_COLUMNS:
            if required and not cell_text(row.get(key)):
                errors.append(import_error(row_number, label, "الحقل إلزامي"))

        department = department_by_key.get(department_key)
        if department_key and not department:
            errors.append(import_error(row_number, "كود الإدارة", "الإدارة غير موجودة"))

        try:
            role = UserRole(role_text)
        except ValueError:
            role = None
            errors.append(import_error(row_number, "الصلاحية", "الصلاحية غير صحيحة"))

        if role == UserRole.INFOSEC:
            errors.append(import_error(row_number, "الصلاحية", "دور أمن المعلومات قديم وغير متاح للاستخدام"))
        if role == UserRole.IT_STAFF and not administrative_section:
            errors.append(import_error(row_number, "كود القسم المختص", "القسم المختص مطلوب لمختص التنفيذ"))
        if administrative_section and administrative_section not in active_sections:
            errors.append(import_error(row_number, "كود القسم المختص", "القسم المختص غير موجود أو غير نشط"))

        if employee_id:
            if employee_id in existing_by_employee or employee_id in seen_employee_ids:
                errors.append(import_error(row_number, "الرقم الوظيفي", "الرقم الوظيفي مكرر"))
            seen_employee_ids.add(employee_id)
        if email:
            if email in existing_emails or email in seen_emails:
                errors.append(import_error(row_number, "البريد الإلكتروني", "البريد الإلكتروني مكرر"))
            seen_emails.add(email)
        if username:
            username_key = username.lower()
            if username_key in existing_usernames or username_key in seen_usernames:
                errors.append(import_error(row_number, "اسم المستخدم", "اسم المستخدم مكرر"))
            seen_usernames.add(username_key)

        try:
            is_active = parse_bool(row.get("is_active"), default=True)
        except ValueError as exc:
            is_active = True
            errors.append(import_error(row_number, "حساب نشط", str(exc)))

        try:
            validate_password_policy(password, policy)
        except HTTPException as exc:
            errors.append(import_error(row_number, "كلمة المرور المؤقتة", str(exc.detail)))

        payload_data = {
            "employee_id": employee_id,
            "username": username,
            "full_name_ar": cell_text(row.get("full_name_ar")),
            "full_name_en": cell_text(row.get("full_name_en")),
            "email": email,
            "mobile": cell_text(row.get("mobile")) or None,
            "password": password,
            "role": role_text,
            "department_id": department.id if department else None,
            "manager_id": None,
            "administrative_section": administrative_section if role == UserRole.IT_STAFF else None,
        }
        try:
            payload = UserCreate.model_validate(payload_data)
            if role:
                ensure_role_assignment_allowed(actor, payload.role)
        except (ValidationError, HTTPException) as exc:
            message = str(exc.detail) if isinstance(exc, HTTPException) else "; ".join(error["msg"] for error in exc.errors())
            errors.append(import_error(row_number, "البيانات", message))
            payload = None

        rows.append({"row_number": row_number, "payload": payload, "manager_employee_id": manager_employee_id, "is_active": is_active})

    new_employee_ids = {item["payload"].employee_id for item in rows if item["payload"]}
    for item in rows:
        payload = item["payload"]
        manager_employee_id = item["manager_employee_id"]
        if not payload or not manager_employee_id:
            continue
        if manager_employee_id not in existing_by_employee and manager_employee_id not in new_employee_ids:
            errors.append(import_error(item["row_number"], "الرقم الوظيفي للمدير المباشر", "المدير غير موجود في النظام أو الملف"))

    if errors:
        raise HTTPException(status_code=422, detail={"message": "لم يتم إنشاء أي مستخدم بسبب وجود أخطاء في الملف", "errors": errors})

    created_users: list[User] = []
    created_by_employee: dict[str, User] = {}
    try:
        for item in rows:
            payload = item["payload"]
            if not payload:
                continue
            user = User(
                employee_id=payload.employee_id,
                username=payload.username,
                full_name_ar=payload.full_name_ar,
                full_name_en=payload.full_name_en,
                email=str(payload.email),
                mobile=payload.mobile,
                hashed_password=get_password_hash(payload.password),
                password_changed_at=datetime.now(timezone.utc),
                role=payload.role,
                administrative_section=payload.administrative_section if payload.role == UserRole.IT_STAFF else None,
                department_id=payload.department_id,
                is_active=item["is_active"],
            )
            db.add(user)
            db.flush()
            created_users.append(user)
            created_by_employee[user.employee_id] = user

        all_managers = {**existing_by_employee, **created_by_employee}
        for item in rows:
            payload = item["payload"]
            manager_employee_id = item["manager_employee_id"]
            if not payload or not manager_employee_id:
                continue
            user = created_by_employee[payload.employee_id]
            manager = all_managers[manager_employee_id]
            payload_with_manager = UserCreate(
                **payload.model_dump(exclude={"manager_id"}),
                manager_id=manager.id,
            )
            validate_user_links(db, payload_with_manager, user_id=user.id)
            user.manager_id = manager.id

        for user in created_users:
            write_audit(db, "user_imported", "user", actor=actor, entity_id=str(user.id), metadata={"email": user.email})
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {"created": len(created_users), "skipped": 0, "errors": []}


def parse_import_workbook(workbook, db: Session, actor: User) -> tuple[list[dict], list[dict]]:
    sheet = workbook["users"] if "users" in workbook.sheetnames else workbook.active
    headers = [IMPORT_HEADER_ALIASES.get(cell_text(value), "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    rows: list[dict] = []
    errors: list[dict] = []
    departments = db.scalars(select(Department)).all()
    department_by_key = {}
    for department in departments:
        for key in [department.id, department.code, department.name_ar, department.name_en]:
            if key is not None and cell_text(key):
                department_by_key[cell_text(key).lower()] = department
    existing_users = db.scalars(select(User)).all()
    existing_employee = {user.employee_id for user in existing_users}
    existing_emails = {user.email.lower() for user in existing_users}
    existing_usernames = {user.username.lower() for user in existing_users if user.username}
    seen_employee: set[str] = set()
    seen_email: set[str] = set()
    seen_username: set[str] = set()
    policy = security_policy(db)
    default_password = temporary_password_from_policy(policy)
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell_text(value) for value in values):
            continue
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]}
        employee_id = cell_text(row.get("employee_id"))
        username = cell_text(row.get("username")) or None
        email = cell_text(row.get("email")).lower()
        department = department_by_key.get(cell_text(row.get("department_code")).lower())
        role_text = cell_text(row.get("role"))
        password = cell_text(row.get("password")) or default_password
        row_errors = []
        for key, label, required in IMPORT_COLUMNS:
            if required and not cell_text(row.get(key)):
                row_errors.append(import_error(row_number, label, "الحقل إلزامي"))
        if employee_id in existing_employee or employee_id in seen_employee:
            row_errors.append(import_error(row_number, "الرقم الوظيفي", "الرقم الوظيفي مكرر"))
        if email in existing_emails or email in seen_email:
            row_errors.append(import_error(row_number, "البريد الإلكتروني", "البريد الإلكتروني مكرر"))
        if username and (username.lower() in existing_usernames or username.lower() in seen_username):
            row_errors.append(import_error(row_number, "اسم المستخدم", "اسم المستخدم مكرر"))
        if not department:
            row_errors.append(import_error(row_number, "كود الإدارة", "الإدارة غير موجودة"))
        try:
            role = UserRole(role_text)
            ensure_role_assignment_allowed(actor, role)
        except Exception:
            role = None
            row_errors.append(import_error(row_number, "الصلاحية", "الصلاحية غير صحيحة"))
        try:
            validate_password_policy(password, policy)
        except HTTPException as exc:
            row_errors.append(import_error(row_number, "كلمة المرور المؤقتة", str(exc.detail)))
        try:
            is_active = parse_bool(row.get("is_active"), default=True)
        except ValueError as exc:
            is_active = True
            row_errors.append(import_error(row_number, "حساب نشط", str(exc)))
        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(
                {
                    "row_number": row_number,
                    "employee_id": employee_id,
                    "username": username,
                    "full_name_ar": cell_text(row.get("full_name_ar")),
                    "full_name_en": cell_text(row.get("full_name_en")),
                    "email": email,
                    "mobile": cell_text(row.get("mobile")) or None,
                    "job_title": cell_text(row.get("job_title")) or None,
                    "password": password,
                    "role": role.value if role else role_text,
                    "department_id": department.id if department else None,
                    "manager_employee_id": cell_text(row.get("manager_employee_id")) or None,
                    "relationship_type": cell_text(row.get("relationship_type")) or "employee",
                    "administrative_section": cell_text(row.get("administrative_section")) or None,
                    "is_active": is_active,
                }
            )
        if employee_id:
            seen_employee.add(employee_id)
        if email:
            seen_email.add(email)
        if username:
            seen_username.add(username.lower())
    return rows, errors


@router.post("/import/validate")
async def validate_users_import(file: UploadFile = File(...), db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="يرجى رفع ملف Excel بصيغة .xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    rows, errors = parse_import_workbook(workbook, db, actor)
    batch = UserImportBatch(
        file_name=file.filename,
        total_rows=len(rows) + len({item["row"] for item in errors}),
        valid_rows=len(rows),
        invalid_rows=len({item["row"] for item in errors}),
        imported_rows=0,
        status="validated" if not errors else "has_errors",
        uploaded_by=actor.id,
        rows_json=rows,
    )
    db.add(batch)
    db.flush()
    for item in errors:
        db.add(UserImportError(batch_id=batch.id, row_number=item["row"], field_name=str(item["field"]), error_message=item["message"]))
    write_audit(db, "user_import_validated", "user", actor=actor, entity_id=str(batch.id), metadata={"valid_rows": len(rows), "invalid_rows": batch.invalid_rows})
    db.commit()
    return {"batch_id": batch.id, "total_rows": batch.total_rows, "valid_rows": batch.valid_rows, "invalid_rows": batch.invalid_rows, "status": batch.status, "errors": errors, "valid_preview": rows[:20]}


@router.post("/import/confirm")
def confirm_users_import(payload: ImportConfirmPayload, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    batch = db.get(UserImportBatch, payload.batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="دفعة الاستيراد غير موجودة")
    if batch.status == "imported":
        raise HTTPException(status_code=409, detail="تم تنفيذ هذه الدفعة مسبقاً")
    if batch.invalid_rows and not payload.import_valid_only:
        raise HTTPException(status_code=422, detail="لا يمكن استيراد الصفوف غير الصالحة")
    created = 0
    by_employee = {user.employee_id: user for user in db.scalars(select(User)).all()}
    created_by_employee: dict[str, User] = {}
    for row in batch.rows_json or []:
        if row["employee_id"] in by_employee:
            continue
        user = User(
            employee_id=row["employee_id"],
            username=row.get("username"),
            full_name_ar=row["full_name_ar"],
            full_name_en=row["full_name_en"],
            email=row["email"],
            mobile=row.get("mobile"),
            job_title=row.get("job_title"),
            hashed_password=get_password_hash(row["password"]),
            role=UserRole(row["role"]),
            department_id=row.get("department_id"),
            relationship_type=row.get("relationship_type") or "employee",
            administrative_section=row.get("administrative_section"),
            is_active=row.get("is_active", True),
            force_password_change=True,
            password_changed_at=datetime.now(timezone.utc),
        )
        db.add(user)
        db.flush()
        created += 1
        created_by_employee[user.employee_id] = user
    all_by_employee = {**by_employee, **created_by_employee}
    for row in batch.rows_json or []:
        manager_employee_id = row.get("manager_employee_id")
        user = created_by_employee.get(row["employee_id"])
        if user and manager_employee_id and manager_employee_id in all_by_employee:
            user.manager_id = all_by_employee[manager_employee_id].id
    batch.imported_rows = created
    batch.status = "imported"
    batch.confirmed_at = datetime.now(timezone.utc)
    write_audit(db, "user_import_confirmed", "user", actor=actor, entity_id=str(batch.id), metadata={"imported_rows": created})
    db.commit()
    return {"imported_rows": created, "batch_id": batch.id}


@router.get("/import/batches")
def import_batches(db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    rows = db.scalars(select(UserImportBatch).options(selectinload(UserImportBatch.uploader)).order_by(UserImportBatch.uploaded_at.desc()).limit(100)).all()
    return [
        {
            "id": row.id,
            "file_name": row.file_name,
            "total_rows": row.total_rows,
            "valid_rows": row.valid_rows,
            "invalid_rows": row.invalid_rows,
            "imported_rows": row.imported_rows,
            "status": row.status,
            "uploaded_by": row.uploader.full_name_ar if row.uploader else "-",
            "uploaded_at": row.uploaded_at,
            "confirmed_at": row.confirmed_at,
        }
        for row in rows
    ]


@router.get("/import/batches/{batch_id}")
def import_batch_details(batch_id: int, db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    batch = db.get(UserImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="دفعة الاستيراد غير موجودة")
    errors = db.scalars(select(UserImportError).where(UserImportError.batch_id == batch.id)).all()
    return {
        "id": batch.id,
        "file_name": batch.file_name,
        "status": batch.status,
        "rows": batch.rows_json or [],
        "errors": [{"row": item.row_number, "field": item.field_name, "message": item.error_message} for item in errors],
    }


@router.get("/overview")
def users_overview(db: Session = Depends(get_db), actor: User = Depends(require_users_screen_view)):
    stmt = select(User)
    if actor.role == UserRole.DIRECT_MANAGER:
        stmt = stmt.where(User.department_id == actor.department_id)
    users = db.scalars(stmt).all()
    departments = db.scalars(select(Department)).all()
    now = datetime.now(timezone.utc)
    admin_roles = {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER, UserRole.EXECUTIVE}
    last_import = db.scalar(select(UserImportBatch).order_by(UserImportBatch.uploaded_at.desc()).limit(1))
    last_permission = db.scalar(
        select(AuditLog).where(AuditLog.action.in_(["user_screen_permissions_updated", "screen_permission_changed", "action_permission_changed", "role_assigned"])).order_by(AuditLog.created_at.desc()).limit(1)
    )
    active_sessions = db.scalar(select(func.count()).select_from(UserSession).where(UserSession.is_active == True, UserSession.revoked_at.is_(None))) or 0
    return {
        "total_users": len(users),
        "active_users": len([user for user in users if user.is_active and not locked_now(user)]),
        "inactive_users": len([user for user in users if not user.is_active]),
        "locked_users": len([user for user in users if locked_now(user)]),
        "without_manager": len([user for user in users if user.role != UserRole.DIRECT_MANAGER and not user.manager_id]),
        "without_department": len([user for user in users if not user.department_id]),
        "admin_users": len([user for user in users if user.role in admin_roles]),
        "last_import_at": last_import.uploaded_at if last_import else None,
        "last_permission_change_at": last_permission.created_at if last_permission else None,
        "active_sessions": active_sessions,
        "users_by_department": [{"label": department.name_ar, "value": len([user for user in users if user.department_id == department.id])} for department in departments],
        "users_by_role": [{"label": role_label(role.value), "value": len([user for user in users if user.role == role])} for role in ACTIVE_USER_ROLES],
        "active_vs_inactive": [
            {"label": "نشط", "value": len([user for user in users if user.is_active and not locked_now(user)])},
            {"label": "غير نشط", "value": len([user for user in users if not user.is_active])},
            {"label": "مقفل", "value": len([user for user in users if locked_now(user)])},
        ],
    }


@router.get("/audit-logs")
def users_audit_logs(
    db: Session = Depends(get_db),
    _: User = Depends(require_users_screen_view),
    action: str | None = Query(default=None),
    user_id: int | None = Query(default=None),
):
    stmt = select(AuditLog).options(selectinload(AuditLog.actor)).where(AuditLog.entity_type.in_(["user", "permission", "role", "delegation", "access_review"]))
    if action:
        stmt = stmt.where(AuditLog.action == action)
    if user_id:
        stmt = stmt.where(AuditLog.entity_id == str(user_id))
    rows = db.scalars(stmt.order_by(AuditLog.created_at.desc()).limit(300)).all()
    return [
        {
            "id": row.id,
            "action": row.action,
            "affected_user_id": row.entity_id if row.entity_type == "user" else None,
            "performed_by": row.actor.full_name_ar if row.actor else "-",
            "performed_by_id": row.actor_id,
            "created_at": row.created_at,
            "ip_address": row.ip_address,
            "old_value": (row.metadata_json or {}).get("old_value"),
            "new_value": (row.metadata_json or {}).get("new_value"),
            "result": (row.metadata_json or {}).get("result", "success"),
            "details": row.metadata_json or {},
        }
        for row in rows
    ]


@router.get("/sessions")
def user_sessions(db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    rows = db.scalars(select(UserSession).options(selectinload(UserSession.user)).order_by(UserSession.login_at.desc()).limit(300)).all()
    return [
        {
            "id": row.id,
            "user_id": row.user_id,
            "user_name": row.user.full_name_ar if row.user else "-",
            "ip_address": row.ip_address,
            "user_agent": row.user_agent,
            "login_at": row.login_at,
            "last_activity_at": row.last_activity_at,
            "revoked_at": row.revoked_at,
            "is_active": row.is_active and not row.revoked_at,
        }
        for row in rows
    ]


@router.get("/login-attempts")
def user_login_attempts(db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    rows = db.scalars(select(UserLoginAttempt).options(selectinload(UserLoginAttempt.user)).order_by(UserLoginAttempt.created_at.desc()).limit(300)).all()
    return [
        {
            "id": row.id,
            "email_or_username": row.email_or_username,
            "user_id": row.user_id,
            "user_name": row.user.full_name_ar if row.user else "-",
            "ip_address": row.ip_address,
            "user_agent": row.user_agent,
            "success": row.success,
            "failure_reason": row.failure_reason,
            "created_at": row.created_at,
        }
        for row in rows
    ]


@router.post("/sessions/{session_id}/revoke")
def revoke_session(session_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    session = db.get(UserSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="الجلسة غير موجودة")
    session.is_active = False
    session.revoked_at = datetime.now(timezone.utc)
    write_audit(db, "sessions_terminated", "user", actor=actor, entity_id=str(session.user_id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"session_id": session_id})
    db.commit()
    return {"status": "revoked"}


@router.post("/sessions/revoke-all")
def revoke_all_sessions(payload: PasswordConfirmPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    if payload.confirmation_text != "REVOKE SESSIONS":
        raise HTTPException(status_code=422, detail="عبارة التأكيد غير صحيحة")
    require_actor_password(actor, payload.admin_password)
    rows = db.scalars(select(UserSession).where(UserSession.is_active == True, UserSession.revoked_at.is_(None))).all()
    now = datetime.now(timezone.utc)
    for row in rows:
        row.is_active = False
        row.revoked_at = now
    write_audit(db, "sessions_terminated", "user", actor=actor, ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"scope": "all", "count": len(rows)})
    db.commit()
    return {"revoked": len(rows)}


@router.get("/organization/tree")
def organization_tree(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    departments = db.scalars(select(Department).order_by(Department.name_ar)).all()
    users = db.scalars(select(User).order_by(User.full_name_ar)).all()
    return [
        {
            "id": department.id,
            "name_ar": department.name_ar,
            "code": department.code,
            "manager": user_display(department.manager),
            "users": [user_display(user) | {"role": user.role, "manager_id": user.manager_id, "is_active": user.is_active} for user in users if user.department_id == department.id],
        }
        for department in departments
    ]


@router.get("/organization/issues")
def organization_issues(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    users = db.scalars(select(User)).all()
    departments = {department.id: department for department in db.scalars(select(Department)).all()}
    issues = []
    for user in users:
        if not user.department_id:
            issues.append({"type": "no_department", "severity": "warning", "user": user_display(user), "message": "مستخدم بدون إدارة"})
        elif user.department_id not in departments or not departments[user.department_id].is_active:
            issues.append({"type": "inactive_department", "severity": "critical", "user": user_display(user), "message": "مستخدم مرتبط بإدارة غير نشطة أو غير موجودة"})
        if user.role not in {UserRole.DIRECT_MANAGER, UserRole.SUPER_ADMIN} and not user.manager_id:
            issues.append({"type": "no_manager", "severity": "warning", "user": user_display(user), "message": "مستخدم بدون مدير مباشر"})
        if user.manager_id == user.id:
            issues.append({"type": "circular_manager", "severity": "critical", "user": user_display(user), "message": "المستخدم لا يمكن أن يكون مدير نفسه"})
    manager_ids = {user.manager_id for user in users if user.manager_id}
    for manager in [user for user in users if user.role in {UserRole.DIRECT_MANAGER, UserRole.DEPARTMENT_MANAGER} and user.id not in manager_ids]:
        issues.append({"type": "manager_without_employees", "severity": "info", "user": user_display(manager), "message": "مدير بدون موظفين"})
    return issues


@router.post("/bulk-assign-department")
def bulk_assign_department(payload: BulkAssignDepartmentPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    department = db.get(Department, payload.department_id)
    if not department or not department.is_active:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة أو غير نشطة")
    count = 0
    for user in db.scalars(select(User).where(User.id.in_(payload.user_ids))).all():
        user.department_id = department.id
        count += 1
    write_audit(db, "bulk_department_assigned", "user", actor=actor, ip_address=client_ip(request), metadata={"department_id": department.id, "user_ids": payload.user_ids})
    db.commit()
    return {"updated": count}


@router.post("/bulk-assign-manager")
def bulk_assign_manager(payload: BulkAssignManagerPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    manager = db.get(User, payload.manager_id)
    if not manager or not manager.is_active:
        raise HTTPException(status_code=404, detail="المدير غير موجود أو غير نشط")
    count = 0
    for user in db.scalars(select(User).where(User.id.in_(payload.user_ids))).all():
        if user.id == manager.id:
            continue
        user.manager_id = manager.id
        count += 1
    write_audit(db, "bulk_manager_assigned", "user", actor=actor, ip_address=client_ip(request), metadata={"manager_id": manager.id, "user_ids": payload.user_ids})
    db.commit()
    return {"updated": count}


@router.get("/delegations")
def list_delegations(db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    rows = db.scalars(select(UserDelegation).options(selectinload(UserDelegation.delegator), selectinload(UserDelegation.delegate)).order_by(UserDelegation.created_at.desc())).all()
    return [
        {
            "id": row.id,
            "delegator_user_id": row.delegator_user_id,
            "delegator_name": row.delegator.full_name_ar if row.delegator else "-",
            "delegate_user_id": row.delegate_user_id,
            "delegate_name": row.delegate.full_name_ar if row.delegate else "-",
            "delegation_scope": row.delegation_scope,
            "start_date": row.start_date,
            "end_date": row.end_date,
            "is_active": row.is_active,
            "reason": row.reason,
            "created_at": row.created_at,
        }
        for row in rows
    ]


@router.get("/delegations/me")
def my_active_delegations(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    rows = db.scalars(
        select(UserDelegation)
        .options(selectinload(UserDelegation.delegator), selectinload(UserDelegation.delegate))
        .where(
            UserDelegation.delegate_user_id == current_user.id,
            UserDelegation.is_active == True,
            UserDelegation.start_date <= now,
            UserDelegation.end_date >= now,
        )
        .order_by(UserDelegation.end_date.asc())
    ).all()
    return [
        {
            "id": row.id,
            "delegator_user_id": row.delegator_user_id,
            "delegator_name": row.delegator.full_name_ar if row.delegator else "-",
            "delegator_role": row.delegator.role if row.delegator else None,
            "delegate_user_id": row.delegate_user_id,
            "delegate_name": row.delegate.full_name_ar if row.delegate else "-",
            "delegation_scope": row.delegation_scope,
            "start_date": row.start_date,
            "end_date": row.end_date,
            "reason": row.reason,
        }
        for row in rows
    ]


@router.post("/delegations")
def create_delegation(payload: DelegationPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    delegator = db.get(User, payload.delegator_user_id)
    delegate = db.get(User, payload.delegate_user_id)
    if not delegator or not delegate or not delegate.is_active:
        raise HTTPException(status_code=422, detail="المفوّض أو البديل غير صحيح")
    if payload.delegator_user_id == payload.delegate_user_id:
        raise HTTPException(status_code=422, detail="لا يمكن تفويض المستخدم لنفسه")
    start_date = normalize_local_datetime(db, payload.start_date)
    end_date = normalize_local_datetime(db, payload.end_date)
    if end_date <= start_date:
        raise HTTPException(status_code=422, detail="تاريخ نهاية التفويض يجب أن يكون بعد تاريخ البداية")
    overlap = db.scalar(
        select(UserDelegation).where(
            UserDelegation.delegator_user_id == payload.delegator_user_id,
            UserDelegation.delegation_scope == payload.delegation_scope,
            UserDelegation.is_active == True,
            UserDelegation.end_date >= start_date,
            UserDelegation.start_date <= end_date,
        )
    )
    if overlap:
        raise HTTPException(status_code=409, detail="يوجد تفويض نشط متداخل لنفس النطاق")
    row = UserDelegation(**payload.model_dump(exclude={"start_date", "end_date"}), start_date=start_date, end_date=end_date, created_by=actor.id)
    db.add(row)
    db.flush()
    write_audit(db, "delegation_created", "delegation", actor=actor, entity_id=str(row.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata=payload.model_dump(mode="json"))
    db.commit()
    return {"id": row.id}


@router.put("/delegations/{delegation_id}")
def update_delegation(delegation_id: int, payload: DelegationPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    row = db.get(UserDelegation, delegation_id)
    if not row:
        raise HTTPException(status_code=404, detail="التفويض غير موجود")
    values = payload.model_dump()
    values["start_date"] = normalize_local_datetime(db, payload.start_date)
    values["end_date"] = normalize_local_datetime(db, payload.end_date)
    if values["end_date"] <= values["start_date"]:
        raise HTTPException(status_code=422, detail="تاريخ نهاية التفويض يجب أن يكون بعد تاريخ البداية")
    for field, value in values.items():
        setattr(row, field, value)
    write_audit(db, "delegation_updated", "delegation", actor=actor, entity_id=str(row.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata=payload.model_dump(mode="json"))
    db.commit()
    return {"id": row.id}


@router.delete("/delegations/{delegation_id}")
def delete_delegation(delegation_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    row = db.get(UserDelegation, delegation_id)
    if not row:
        raise HTTPException(status_code=404, detail="التفويض غير موجود")
    metadata = {
        "delegator_user_id": row.delegator_user_id,
        "delegate_user_id": row.delegate_user_id,
        "delegation_scope": row.delegation_scope,
        "start_date": row.start_date.isoformat() if row.start_date else None,
        "end_date": row.end_date.isoformat() if row.end_date else None,
        "reason": row.reason,
    }
    write_audit(db, "delegation_deleted", "delegation", actor=actor, entity_id=str(row.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata=metadata)
    db.delete(row)
    db.commit()
    return {"status": "deleted"}


def access_review_issues(db: Session) -> list[dict]:
    users = db.scalars(select(User)).all()
    direct_override_user_ids = {
        int(row.setting_key)
        for row in db.scalars(select(PortalSetting).where(PortalSetting.category == "screen_permissions")).all()
        if str(row.setting_key).isdigit()
    }
    action_override_user_ids = {row.user_id for row in db.scalars(select(ActionPermission).where(ActionPermission.user_id.is_not(None))).all()}
    issues = []
    threshold = datetime.now(timezone.utc) - timedelta(days=30)
    for user in users:
        if user.role == UserRole.SUPER_ADMIN:
            issues.append({"user": user_display(user), "issue_type": "super_admin", "description": "مستخدم بصلاحية مدير النظام", "status": "action_required"})
        if not user.department_id:
            issues.append({"user": user_display(user), "issue_type": "no_department", "description": "مستخدم بدون إدارة", "status": "pending"})
        if user.role not in {UserRole.DIRECT_MANAGER, UserRole.SUPER_ADMIN} and not user.manager_id:
            issues.append({"user": user_display(user), "issue_type": "no_manager", "description": "مستخدم بدون مدير مباشر", "status": "pending"})
        if user.last_login_at and (user.last_login_at.replace(tzinfo=timezone.utc) if user.last_login_at.tzinfo is None else user.last_login_at) < threshold:
            issues.append({"user": user_display(user), "issue_type": "inactive_login", "description": "لم يسجل الدخول منذ أكثر من 30 يوم", "status": "pending"})
        if not user.is_active and (user.id in direct_override_user_ids or user.id in action_override_user_ids):
            issues.append({"user": user_display(user), "issue_type": "inactive_with_permissions", "description": "مستخدم غير نشط لديه صلاحيات مباشرة", "status": "action_required"})
        if user.id in direct_override_user_ids or user.id in action_override_user_ids:
            issues.append({"user": user_display(user), "issue_type": "direct_overrides", "description": "لديه صلاحيات مباشرة مخصصة", "status": "pending"})
    return issues


@router.get("/access-review")
def get_access_review(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    latest = db.scalar(select(AccessReview).order_by(AccessReview.created_at.desc()).limit(1))
    saved_items = []
    if latest:
        rows = db.scalars(
            select(AccessReviewItem)
            .options(selectinload(AccessReviewItem.user))
            .where(AccessReviewItem.review_id == latest.id)
            .order_by(AccessReviewItem.status, AccessReviewItem.issue_type)
        ).all()
        saved_items = [
            {
                "id": row.id,
                "review_id": row.review_id,
                "user": user_display(row.user),
                "issue_type": row.issue_type,
                "description": row.description,
                "status": row.status,
                "reviewed_by": row.reviewed_by,
                "reviewed_at": row.reviewed_at,
            }
            for row in rows
        ]
    return {
        "latest_review": {
            "id": latest.id,
            "review_name": latest.review_name,
            "status": latest.status,
            "created_at": latest.created_at,
            "completed_at": latest.completed_at,
        }
        if latest
        else None,
        "items": access_review_issues(db),
        "saved_items": saved_items,
    }


@router.post("/access-review")
def create_access_review(request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    review = AccessReview(review_name=f"مراجعة صلاحيات {datetime.now().strftime('%Y-%m-%d')}", status="pending", created_by=actor.id)
    db.add(review)
    db.flush()
    count = 0
    for item in access_review_issues(db):
        db.add(AccessReviewItem(review_id=review.id, user_id=item["user"]["id"], issue_type=item["issue_type"], description=item["description"], status=item["status"]))
        count += 1
    write_audit(db, "access_review_created", "access_review", actor=actor, entity_id=str(review.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"items": count})
    db.commit()
    return {"review_id": review.id, "items": count}


@router.post("/access-review/{review_id}/complete")
def complete_access_review(review_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    review = db.get(AccessReview, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="المراجعة غير موجودة")
    review.status = "reviewed"
    review.completed_at = datetime.now(timezone.utc)
    write_audit(db, "access_review_completed", "access_review", actor=actor, entity_id=str(review.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    return {"status": review.status}


@router.post("/access-review/items/{item_id}/mark-reviewed")
def mark_access_review_item(item_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    item = db.get(AccessReviewItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="البند غير موجود")
    item.status = "reviewed"
    item.reviewed_by = actor.id
    item.reviewed_at = datetime.now(timezone.utc)
    write_audit(db, "access_review_item_reviewed", "access_review", actor=actor, entity_id=str(item.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    return {"status": item.status}


@router.get("", response_model=list[UserRead])
def list_users(db: Session = Depends(get_db), actor: User = Depends(require_users_screen_view)):
    stmt = select(User).order_by(User.full_name_ar)
    if actor.role == UserRole.DIRECT_MANAGER:
        stmt = stmt.where(User.department_id == actor.department_id)
    return db.scalars(stmt).all()


@router.get("/screen-permissions/me", response_model=ScreenPermissionsRead)
def my_screen_permissions(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    screens = read_user_screens(db, current_user)
    if has_active_approval_delegation(db, current_user) and "approvals" not in screens:
        screens.append("approvals")
    return {"screens": screens, "available_screens": available_screens_for_user(current_user)}


@router.get("/{user_id}")
def get_user_details(user_id: int, db: Session = Depends(get_db), actor: User = Depends(require_users_screen_view)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role == UserRole.DIRECT_MANAGER and user.department_id != actor.department_id:
        raise HTTPException(status_code=403, detail="لا يمكنك عرض مستخدم خارج إدارتك")
    audit_logs = db.scalars(select(AuditLog).options(selectinload(AuditLog.actor)).where(AuditLog.entity_type == "user", AuditLog.entity_id == str(user.id)).order_by(AuditLog.created_at.desc()).limit(20)).all()
    sessions = db.scalars(select(UserSession).where(UserSession.user_id == user.id).order_by(UserSession.login_at.desc()).limit(10)).all()
    return {
        "user": UserRead.model_validate(user).model_dump(mode="json"),
        "manager": user_display(user.manager),
        "screen_permissions": read_user_screens(db, user),
        "action_permissions": read_action_permissions(db, user_id=user.id),
        "sessions": [
            {"id": session.id, "ip_address": session.ip_address, "user_agent": session.user_agent, "login_at": session.login_at, "last_activity_at": session.last_activity_at, "is_active": session.is_active and not session.revoked_at}
            for session in sessions
        ],
        "recent_audit_logs": [
            {"id": log.id, "action": log.action, "performed_by": log.actor.full_name_ar if log.actor else "-", "created_at": log.created_at, "details": log.metadata_json or {}}
            for log in audit_logs
        ],
    }


@router.get("/{user_id}/audit-logs")
def get_user_audit_logs(user_id: int, db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    rows = db.scalars(select(AuditLog).options(selectinload(AuditLog.actor)).where(AuditLog.entity_type == "user", AuditLog.entity_id == str(user_id)).order_by(AuditLog.created_at.desc()).limit(200)).all()
    return [
        {"id": row.id, "action": row.action, "performed_by": row.actor.full_name_ar if row.actor else "-", "created_at": row.created_at, "ip_address": row.ip_address, "details": row.metadata_json or {}}
        for row in rows
    ]


@router.get("/{user_id}/screen-permissions", response_model=ScreenPermissionsRead)
def get_user_screen_permissions(user_id: int, db: Session = Depends(get_db), _: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {"screens": read_user_screens(db, user), "available_screens": available_screens_for_user(user)}


@router.put("/{user_id}/screen-permissions", response_model=ScreenPermissionsRead)
def update_user_screen_permissions(user_id: int, payload: ScreenPermissionsPayload, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role != UserRole.SUPER_ADMIN and user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can update Super Admin permissions")
    save_user_screens(db, user, payload.screens, actor)
    write_audit(db, "user_screen_permissions_updated", "user", actor=actor, entity_id=str(user.id), metadata={"screens": payload.screens})
    db.commit()
    return {"screens": read_user_screens(db, user), "available_screens": available_screens_for_user(user)}


@router.post("", response_model=UserRead)
def create_user(payload: UserCreate, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_users_screen_create)):
    actor_can_manage_users = permission_level_allows(effective_screen_permission_level(db, actor, "users"), "manage")
    if actor.role not in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER} and not actor_can_manage_users and payload.role != UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="صلاحية إضافة المستخدمين لا تسمح بإنشاء أدوار إدارية")
    ensure_role_assignment_allowed(actor, payload.role)
    validate_user_links(db, payload)
    ensure_user_unique(db, payload)
    policy = security_policy(db)
    password = payload.password or temporary_password_from_policy(policy)
    validate_password_policy(password, policy)
    user = User(
        employee_id=payload.employee_id,
        username=payload.username,
        full_name_ar=payload.full_name_ar,
        full_name_en=payload.full_name_en,
        email=str(payload.email),
        mobile=payload.mobile,
        job_title=payload.job_title,
        hashed_password=get_password_hash(password),
        password_changed_at=datetime.now(timezone.utc),
        role=payload.role,
        role_id=payload.role_id,
        administrative_section=payload.administrative_section if payload.role == UserRole.IT_STAFF else None,
        specialized_section_id=payload.specialized_section_id,
        relationship_type=payload.relationship_type or ("direct_manager" if payload.role == UserRole.DIRECT_MANAGER else "employee"),
        department_id=payload.department_id,
        manager_id=payload.manager_id,
        force_password_change=payload.force_password_change,
        password_expires_at=payload.password_expires_at,
        allowed_login_from_ip=payload.allowed_login_from_ip,
        notes=payload.notes,
    )
    db.add(user)
    db.flush()
    write_audit(db, "user_created", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"email": user.email, "new_value": {"role": user.role, "department_id": user.department_id}})
    db.commit()
    db.refresh(user)
    return user


@router.put("/{user_id}", response_model=UserRead)
def update_user(user_id: int, payload: UserUpdate, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_users_screen_edit)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role != UserRole.SUPER_ADMIN and user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can update Super Admin users")
    actor_can_manage_users = permission_level_allows(effective_screen_permission_level(db, actor, "users"), "manage")
    if actor.role not in {UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER} and not actor_can_manage_users:
        if user.role == UserRole.DEPARTMENT_MANAGER:
            raise HTTPException(status_code=403, detail="لا يمكن تعديل مستخدم إداري بهذه الصلاحية")
        if payload.role != user.role:
            raise HTTPException(status_code=403, detail="تغيير دور المستخدم يتطلب صلاحية إدارة المستخدمين")
    ensure_not_last_super_admin(db, user, next_role=payload.role, next_active=payload.is_active)
    ensure_role_assignment_allowed(actor, payload.role)
    validate_user_links(db, payload, user_id=user_id)
    ensure_user_unique(db, payload, user_id=user_id)
    payload_data = payload.model_dump()
    if payload.role != UserRole.IT_STAFF:
        payload_data["administrative_section"] = None
    old_value = {"role": user.role, "department_id": user.department_id, "manager_id": user.manager_id, "is_active": user.is_active}
    for field, value in payload_data.items():
        if field == "email":
            value = str(value)
        setattr(user, field, value)
    user.relationship_type = user.relationship_type or ("direct_manager" if user.role == UserRole.DIRECT_MANAGER else "employee")
    write_audit(db, "user_updated", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"old_value": old_value, "new_value": {"role": user.role, "department_id": user.department_id, "manager_id": user.manager_id, "is_active": user.is_active}})
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/disable", response_model=UserRead)
def disable_user(user_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role != UserRole.SUPER_ADMIN and user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can disable Super Admin users")
    ensure_not_last_super_admin(db, user, next_active=False)
    user.is_active = False
    write_audit(db, "user_disabled", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/enable", response_model=UserRead)
def enable_user(user_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = True
    write_audit(db, "user_enabled", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/lock", response_model=UserRead)
def lock_user(user_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role != UserRole.SUPER_ADMIN and user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can lock Super Admin users")
    user.is_locked = True
    user.locked_until = datetime.now(timezone.utc) + timedelta(days=3650)
    write_audit(db, "user_locked", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/unlock", response_model=UserRead)
def unlock_user(user_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_locked = False
    user.locked_until = None
    user.failed_login_attempts = 0
    write_audit(db, "user_unlocked", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/terminate-sessions")
def terminate_user_sessions(user_id: int, payload: PasswordConfirmPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    require_actor_password(actor, payload.admin_password)
    rows = db.scalars(select(UserSession).where(UserSession.user_id == user.id, UserSession.is_active == True, UserSession.revoked_at.is_(None))).all()
    now = datetime.now(timezone.utc)
    for row in rows:
        row.is_active = False
        row.revoked_at = now
    write_audit(db, "sessions_terminated", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"count": len(rows)})
    db.commit()
    return {"revoked": len(rows)}


@router.post("/{user_id}/reset-password", status_code=status.HTTP_204_NO_CONTENT)
def reset_user_password(user_id: int, payload: PasswordReset, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if actor.role != UserRole.SUPER_ADMIN and user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only Super Admin can reset Super Admin passwords")
    require_actor_password(actor, payload.admin_password)
    policy = security_policy(db)
    password = payload.password or temporary_password_from_policy(policy)
    validate_password_policy(password, policy)
    user.hashed_password = get_password_hash(password)
    user.password_changed_at = datetime.now(timezone.utc)
    user.force_password_change = True
    user.failed_login_attempts = 0
    user.locked_until = None
    user.is_locked = False
    write_audit(db, "password_reset", "user", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request))
    db.commit()


roles_router = APIRouter(prefix="/roles", tags=["Roles"])


@roles_router.get("")
def list_roles(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    counts = dict(db.execute(select(User.role, func.count()).group_by(User.role)).all())
    roles = db.scalars(select(Role).order_by(Role.label_ar)).all()
    result = []
    for role in roles:
        if is_hidden_legacy_role(role):
            continue
        item = role_to_dict(role)
        item["users_count"] = counts.get(role.code or role.name, 0)
        result.append(item)
    return result


@roles_router.post("")
def create_role(payload: RolePayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    if is_hidden_legacy_role_code(payload.code):
        raise HTTPException(status_code=422, detail="لا يمكن استخدام كود دور قديم")
    existing = role_by_code(db, payload.code)
    if existing:
        raise HTTPException(status_code=409, detail="كود الدور مستخدم من قبل")
    role = Role(name=payload.code, label_ar=payload.name_ar, name_ar=payload.name_ar, name_en=payload.name_en or payload.code, code=payload.code, description=payload.description, is_system_role=False, is_active=payload.is_active)
    db.add(role)
    db.flush()
    write_audit(db, "role_created", "role", actor=actor, entity_id=str(role.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"new_value": role_to_dict(role)})
    db.commit()
    db.refresh(role)
    return role_to_dict(role)


@roles_router.put("/{role_id}")
def update_role(role_id: int, payload: RolePayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    if is_hidden_legacy_role(role) or is_hidden_legacy_role_code(payload.code):
        raise HTTPException(status_code=422, detail="لا يمكن تعديل أو إعادة استخدام دور قديم")
    duplicate = db.scalar(select(Role).where(Role.id != role_id, or_(Role.code == payload.code, Role.name == payload.code)))
    if duplicate:
        raise HTTPException(status_code=409, detail="كود الدور مستخدم من قبل")
    old_value = role_to_dict(role)
    active_users = db.scalar(select(func.count()).select_from(User).where(User.role == (role.code or role.name), User.is_active == True)) or 0
    if active_users and not payload.is_active:
        raise HTTPException(status_code=409, detail="لا يمكن تعطيل دور مرتبط بمستخدمين نشطين")
    role.label_ar = payload.name_ar
    role.name_ar = payload.name_ar
    role.name_en = payload.name_en or payload.code
    if not role.is_system_role:
        role.name = payload.code
        role.code = payload.code
    role.description = payload.description
    role.is_active = payload.is_active
    write_audit(db, "role_updated", "role", actor=actor, entity_id=str(role.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"old_value": old_value, "new_value": role_to_dict(role)})
    db.commit()
    db.refresh(role)
    return role_to_dict(role)


@roles_router.patch("/{role_id}/status")
def update_role_status(role_id: int, payload: RolePayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    if is_hidden_legacy_role(role):
        raise HTTPException(status_code=422, detail="لا يمكن تفعيل دور قديم")
    active_users = db.scalar(select(func.count()).select_from(User).where(User.role == (role.code or role.name), User.is_active == True)) or 0
    if active_users and not payload.is_active:
        raise HTTPException(status_code=409, detail="لا يمكن تعطيل دور مرتبط بمستخدمين نشطين")
    role.is_active = payload.is_active
    write_audit(db, "role_status_changed", "role", actor=actor, entity_id=str(role.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"new_value": {"is_active": role.is_active}})
    db.commit()
    return role_to_dict(role)


@roles_router.post("/{role_id}/clone")
def clone_role(role_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    if is_hidden_legacy_role(role):
        raise HTTPException(status_code=422, detail="لا يمكن استنساخ دور قديم")
    base_code = f"{role.code or role.name}_copy"
    code = base_code
    counter = 1
    while role_by_code(db, code):
        counter += 1
        code = f"{base_code}_{counter}"
    clone = Role(name=code, label_ar=f"نسخة من {role.name_ar or role.label_ar}", name_ar=f"نسخة من {role.name_ar or role.label_ar}", name_en=f"Copy of {role.name_en or role.name}", code=code, description=role.description, is_system_role=False, is_active=True)
    db.add(clone)
    db.flush()
    save_role_screens(db, clone, read_role_screens(db, role.code or role.name), actor)
    for action, allowed in read_action_permissions(db, role_id=role.id).items():
        set_action_permissions(db, {action: allowed}, actor, role_id=clone.id)
    write_audit(db, "role_cloned", "role", actor=actor, entity_id=str(clone.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"source_role_id": role.id})
    db.commit()
    return role_to_dict(clone)


@roles_router.delete("/{role_id}")
def delete_role(role_id: int, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    if role.is_system_role:
        raise HTTPException(status_code=409, detail="لا يمكن حذف دور نظامي. يمكن تعديل أو تعطيل الأدوار المخصصة فقط.")

    role_code = role.code or role.name
    linked_users = db.scalar(
        select(func.count())
        .select_from(User)
        .where(or_(User.role == role_code, User.role == role.name, User.role_id == role.id))
    ) or 0
    if linked_users:
        raise HTTPException(status_code=409, detail="لا يمكن حذف الدور لأنه مرتبط بمستخدمين. انقل المستخدمين إلى دور آخر أولاً.")

    workflow_steps = db.scalar(
        select(func.count()).select_from(WorkflowTemplateStep).where(WorkflowTemplateStep.approver_role_id == role.id)
    ) or 0
    if workflow_steps:
        raise HTTPException(status_code=409, detail="لا يمكن حذف الدور لأنه مستخدم في مسارات الموافقة.")

    request_steps = db.scalar(
        select(func.count()).select_from(RequestApprovalStep).where(RequestApprovalStep.approver_role_id == role.id)
    ) or 0
    if request_steps:
        raise HTTPException(status_code=409, detail="لا يمكن حذف الدور لأنه موجود في طلبات سابقة.")

    old_value = role_to_dict(role)
    db.execute(delete(ScreenPermission).where(ScreenPermission.role_id == role.id))
    db.execute(delete(ActionPermission).where(ActionPermission.role_id == role.id))
    db.execute(delete(AIFeaturePermission).where(AIFeaturePermission.role_id == role.id))
    db.delete(role)
    write_audit(db, "role_deleted", "role", actor=actor, entity_id=str(role_id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"old_value": old_value})
    db.commit()
    return {"deleted": True}


@roles_router.get("/{role_id}/users")
def users_in_role(role_id: int, db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    rows = db.scalars(select(User).where(User.role == (role.code or role.name)).order_by(User.full_name_ar)).all()
    return [UserRead.model_validate(user).model_dump(mode="json") for user in rows]


permissions_router = APIRouter(prefix="/permissions", tags=["Permissions"])


@permissions_router.get("/screens")
def screen_permissions_matrix(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    roles = [role for role in db.scalars(select(Role).order_by(Role.label_ar)).all() if not is_hidden_legacy_role(role)]
    users = db.scalars(select(User).order_by(User.full_name_ar)).all()
    return {
        "screens": SCREEN_DEFINITIONS,
        "levels": PERMISSION_LEVELS,
        "roles": [
            {
                "id": role.id,
                "code": role.code or role.name,
                "name_ar": role.name_ar or role.label_ar,
                "screens": read_role_screens(db, role.code or role.name),
                "permissions": read_screen_permission_levels(db, role_id=role.id, fallback_screens=read_role_screens(db, role.code or role.name)),
            }
            for role in roles
        ],
        "users": [
            {
                "id": user.id,
                "name_ar": user.full_name_ar,
                "username": user.username,
                "employee_id": user.employee_id,
                "email": user.email,
                "role": user.role,
                "screens": read_user_screens(db, user),
                "permissions": read_screen_permission_levels(db, user_id=user.id, fallback_screens=read_user_screens(db, user)),
            }
            for user in users
        ],
    }


@permissions_router.get("/screens/effective/{user_id}")
def effective_screen_permissions(user_id: int, db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    screens = read_user_screens(db, user)
    return {"user_id": user.id, "screens": screens, "permissions": read_screen_permission_levels(db, user_id=user.id, fallback_screens=screens), "available_screens": available_screens_for_user(user)}


@permissions_router.put("/screens/role/{role_id}")
def update_role_screen_permissions(role_id: int, payload: PermissionLevelPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    screens = [code for code, level in payload.permissions.items() if level != "no_access" and code in ALL_SCREEN_KEYS]
    save_role_screens(db, role, screens, actor)
    set_screen_permission_levels(db, payload.permissions, role_id=role.id)
    write_audit(db, "screen_permission_changed", "permission", actor=actor, entity_id=str(role.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"subject": "role", "permissions": payload.permissions})
    db.commit()
    saved_screens = read_role_screens(db, role.code or role.name)
    return {"screens": saved_screens, "permissions": read_screen_permission_levels(db, role_id=role.id, fallback_screens=saved_screens)}


@permissions_router.put("/screens/user/{user_id}")
def update_user_screen_permission_matrix(user_id: int, payload: PermissionLevelPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    screens = [code for code, level in payload.permissions.items() if level != "no_access" and code in ALL_SCREEN_KEYS]
    save_user_screens(db, user, screens, actor)
    set_screen_permission_levels(db, payload.permissions, user_id=user.id)
    write_audit(db, "screen_permission_changed", "permission", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"subject": "user", "permissions": payload.permissions})
    db.commit()
    saved_screens = read_user_screens(db, user)
    return {"screens": saved_screens, "permissions": read_screen_permission_levels(db, user_id=user.id, fallback_screens=saved_screens)}


@permissions_router.post("/screens/copy")
def copy_screen_permissions(payload: dict, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    source_user = db.get(User, payload.get("source_user_id")) if payload.get("source_user_id") else None
    target_user = db.get(User, payload.get("target_user_id")) if payload.get("target_user_id") else None
    if not source_user or not target_user:
        raise HTTPException(status_code=404, detail="المستخدم المصدر أو الهدف غير موجود")
    screens = read_user_screens(db, source_user)
    save_user_screens(db, target_user, screens, actor)
    write_audit(db, "screen_permission_changed", "permission", actor=actor, entity_id=str(target_user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"copied_from_user_id": source_user.id, "screens": screens})
    db.commit()
    return {"screens": screens}


@permissions_router.get("/actions")
def action_permissions_matrix(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    roles = [role for role in db.scalars(select(Role).order_by(Role.label_ar)).all() if not is_hidden_legacy_role(role)]
    users = db.scalars(select(User).order_by(User.full_name_ar)).all()
    return {
        "actions": ACTION_DEFINITIONS,
        "roles": [{"id": role.id, "code": role.code or role.name, "name_ar": role.name_ar or role.label_ar, "permissions": read_action_permissions(db, role_id=role.id)} for role in roles],
        "users": [
            {
                "id": user.id,
                "name_ar": user.full_name_ar,
                "username": user.username,
                "employee_id": user.employee_id,
                "email": user.email,
                "role": user.role,
                "permissions": read_action_permissions(db, user_id=user.id),
            }
            for user in users
        ],
    }


@permissions_router.get("/actions/effective/{user_id}")
def effective_action_permissions(user_id: int, db: Session = Depends(get_db), _: User = Depends(require_users_screen_view)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    role = role_by_code(db, user.role)
    result = read_action_permissions(db, role_id=role.id) if role else {}
    result.update(read_action_permissions(db, user_id=user.id))
    return {"user_id": user.id, "permissions": result}


@permissions_router.put("/actions/role/{role_id}")
def update_role_actions(role_id: int, payload: ActionPermissionPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    if any(item["dangerous"] and payload.permissions.get(item["code"]) for item in ACTION_DEFINITIONS) and payload.confirmation_text != "CONFIRM PERMISSIONS":
        raise HTTPException(status_code=422, detail="الصلاحيات الخطرة تتطلب عبارة التأكيد CONFIRM PERMISSIONS")
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    set_action_permissions(db, payload.permissions, actor, role_id=role.id)
    write_audit(db, "action_permission_changed", "permission", actor=actor, entity_id=str(role.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"subject": "role", "new_value": payload.permissions})
    db.commit()
    return {"permissions": read_action_permissions(db, role_id=role.id)}


@permissions_router.put("/actions/user/{user_id}")
def update_user_actions(user_id: int, payload: ActionPermissionPayload, request: Request, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN))):
    if any(item["dangerous"] and payload.permissions.get(item["code"]) for item in ACTION_DEFINITIONS) and payload.confirmation_text != "CONFIRM PERMISSIONS":
        raise HTTPException(status_code=422, detail="الصلاحيات الخطرة تتطلب عبارة التأكيد CONFIRM PERMISSIONS")
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    set_action_permissions(db, payload.permissions, actor, user_id=user.id)
    write_audit(db, "action_permission_changed", "permission", actor=actor, entity_id=str(user.id), ip_address=client_ip(request), user_agent=request_user_agent(request), metadata={"subject": "user", "new_value": payload.permissions})
    db.commit()
    return {"permissions": read_action_permissions(db, user_id=user.id)}


departments_router = APIRouter(prefix="/departments", tags=["Departments"])


@departments_router.get("", response_model=list[DepartmentRead])
def list_departments(db: Session = Depends(get_db), _: User = Depends(require_users_screen_view), search: str | None = Query(default=None)):
    stmt = select(Department).order_by(Department.name_ar)
    if search:
        stmt = stmt.where(Department.name_ar.ilike(f"%{search}%") | Department.name_en.ilike(f"%{search}%"))
    return db.scalars(stmt).all()


@departments_router.post("", response_model=DepartmentRead, status_code=status.HTTP_201_CREATED)
def create_department(payload: SettingsDepartmentCreate, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    validate_department_manager(db, payload.manager_id)
    department = Department(
        name_ar=payload.name_ar,
        name_en=payload.name_en,
        code=payload.code,
        manager_id=payload.manager_id,
        is_active=payload.is_active,
    )
    db.add(department)
    db.flush()
    write_audit(db, "department_created", "department", actor=actor, entity_id=str(department.id))
    db.commit()
    db.refresh(department)
    return department


@departments_router.put("/{department_id}", response_model=DepartmentRead)
def update_department(department_id: int, payload: SettingsDepartmentCreate, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    department = db.get(Department, department_id)
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    validate_department_manager(db, payload.manager_id)
    for field, value in payload.model_dump().items():
        setattr(department, field, value)
    write_audit(db, "department_updated", "department", actor=actor, entity_id=str(department.id))
    db.commit()
    db.refresh(department)
    return department


@departments_router.delete("/{department_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_department(department_id: int, db: Session = Depends(get_db), actor: User = Depends(require_roles(UserRole.SUPER_ADMIN, UserRole.DEPARTMENT_MANAGER))):
    department = db.get(Department, department_id)
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    db.delete(department)
    write_audit(db, "department_deleted", "department", actor=actor, entity_id=str(department_id))
    db.commit()


def validate_department_manager(db: Session, manager_id: int | None) -> None:
    if not manager_id:
        return
    manager = db.get(User, manager_id)
    if not manager or not manager.is_active or manager.is_locked:
        raise HTTPException(status_code=422, detail="مدير الإدارة المحدد غير موجود أو غير نشط")
    if manager.role not in {UserRole.DIRECT_MANAGER, UserRole.DEPARTMENT_MANAGER, UserRole.EXECUTIVE, UserRole.SUPER_ADMIN}:
        raise HTTPException(status_code=422, detail="مدير الإدارة يجب أن يكون مستخدماً بصلاحية إدارية أو مدير مباشر")
